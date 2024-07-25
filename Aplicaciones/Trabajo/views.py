from django.shortcuts import render, redirect, get_object_or_404
from .models import fruta, SalidaStock,Usuario,congeladoP,recep, Nota,Rol,ventas, Maquila,alm,proTer,alfrut,frigoe, despatado,producto, proveedores, lavado,congelado,quebrado,Desperdicio, concervacion, ordenes, Usuario, compras, embalaje, paleta,materia,fisicoQ
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth import login as auth_login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt
import logging
from django.db.models import Sum,F, FloatField
from django.db.models.signals import post_save
from django.dispatch import receiver
from decimal import Decimal,DivisionUndefined
from .utils import is_admin, is_admin_prod,is_admin_recep,is_admin_cal
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from .models import fruta
import openpyxl
import datetime
from datetime import datetime ,timedelta,date,time
from openpyxl.styles import Protection, NamedStyle
from django.utils import timezone
from django.template.defaultfilters import default_if_none
from collections import defaultdict
from django.template.loader import render_to_string
from django.utils.datastructures import MultiValueDictKeyError
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models.functions import TruncDate
from django.conf import settings
from django.db import connection


#------------------------------------Conceptos gnerales----------------------------------- 
def get_last_activity():
    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT usuario_id, last_activity_hour, DATE(last_activity_day) as last_activity_day
            FROM (
                SELECT usuario_id, hora as last_activity_hour, fecha as last_activity_day FROM Trabajo_fruta 
                UNION ALL
                SELECT usuario_id, hora as last_activity_hour, horaEnvio as last_activity_day FROM Trabajo_lavado
                UNION ALL
                SELECT usuario_id, hora as last_activity_hour, horaEnvio as last_activity_day FROM Trabajo_congelado
                UNION ALL
                SELECT usuario_id, hora as last_activity_hour, horaEnvio as last_activity_day FROM Trabajo_quebrado
                UNION ALL
                SELECT nombre_id as usuario_id, hora as last_activity_hour, fecha as last_activity_day FROM Trabajo_concervacion
                UNION ALL
                SELECT usuario_id, hora as last_activity_hour, horaEnvio as last_activity_day FROM Trabajo_maquila
                UNION ALL
                SELECT usuario_id, hora as last_activity_hour, fecha as last_activity_day FROM Trabajo_recep
                UNION ALL
                SELECT usuario_id, hora as last_activity_hour, fecha as last_activity_day FROM Trabajo_fisicoQ
            ) as combined
            ORDER BY usuario_id, last_activity_day DESC, last_activity_hour DESC;
        ''')
        rows = cursor.fetchall()
    return rows

@user_passes_test(is_admin, login_url='/error/' )
def inicio(request):
     # Obtener los IDs de frutas que están en proceso de lavado
    frutas_lavadas_ids = lavado.objects.values_list('lote_id', flat=True)
    print(frutas_lavadas_ids)

    # Filtrar las frutas que no están en proceso de lavado y agrupar por nombre
    frutas_disponibles = fruta.objects.exclude(
        movimiento__in=frutas_lavadas_ids
    ).values('nombre__nombre').annotate(
        saldo_disponible=Sum('kgNetos')
    ).order_by('nombre__nombre')

    last_activities = get_last_activity()

    if not last_activities:
        usuarios_activos = []
    else:
        usuarios = Usuario.objects.filter(id__in=[item[0] for item in last_activities])
        usuarios_activos = [
            {
                'username': usuario.username,
                'rol': usuario.rol,
                'last_activity_hour': next(item[1] for item in last_activities if item[0] == usuario.id),
                'last_activity_day': next(item[2] for item in last_activities if item[0] == usuario.id)
            }
            for usuario in usuarios
        ]
    # Obtener la fecha del filtro, o usar hoy por defecto
    fecha_str = request.GET.get('fecha', datetime.today().strftime('%Y-%m-%d'))
    fecha = datetime.strptime(fecha_str, '%Y-%m-%d')
    start_of_day = datetime.combine(fecha, datetime.min.time())
    end_of_day = start_of_day + timedelta(days=1)
    
    # Obtener el ID de la fruta del filtro, o usar None por defecto
    fruta_id = request.GET.get('fruta')
    
    # Filtrar por día y, si se especifica, por fruta
    lavado_agrupado = lavado.objects.filter(horaEnvio__gte=start_of_day, horaEnvio__lt=end_of_day)
    if fruta_id:
        lavado_agrupado = lavado_agrupado.filter(fruta_id=fruta_id).first
    lavado_agrupado = lavado_agrupado.values('nombre__nombre').annotate(saldo_lavado=Sum('kgNetos')).order_by('nombre__nombre')
    
    congelado_agrupado = congelado.objects.filter(horaEnvio__gte=start_of_day, horaEnvio__lt=end_of_day)
    if fruta_id:
        congelado_agrupado = congelado_agrupado.filter(fruta_id=fruta_id)
    congelado_agrupado = congelado_agrupado.values('nombre__nombre').annotate(saldo_congelado=Sum('kgNetos')).order_by('nombre__nombre')
    
    quebrado_agrupado = quebrado.objects.filter(horaEnvio__gte=start_of_day, horaEnvio__lt=end_of_day)
    if fruta_id:
        quebrado_agrupado = quebrado_agrupado.filter(fruta_id=fruta_id)
    quebrado_agrupado = quebrado_agrupado.values('nombre__nombre').annotate(saldo_quebrado=Sum('kgNetos')).order_by('nombre__nombre')
    
    pt_agrupado = concervacion.objects.filter(fecha__gte=start_of_day, fecha__lt=end_of_day)
    if fruta_id:
        pt_agrupado = pt_agrupado.filter(fruta_id=fruta_id)
    pt_agrupado = pt_agrupado.values('nombre__nombre').annotate(saldo_pt=Sum('kgT')).order_by('nombre__nombre')
    
    # Obtener todas las frutas para el dropdown del filtro
    frutas = fruta.objects.all()
    
    context = {
        'lavado_agrupado': lavado_agrupado,
        'congelado_agrupado': congelado_agrupado,
        'quebrado_agrupado': quebrado_agrupado,
        'pt_agrupado': pt_agrupado,
        'frutas': frutas,
        'frutas_disponibles': frutas_disponibles,
        'usuarios_activos': usuarios_activos,
        'usuarios_activos':usuarios_activos
    }
    return render(request, 'inicio.html', context)

def error(request):
    return render(request, 'Error.html')

def intro(request):

    return render(request, 'login.html')

def login(request):
    return render(request, 'login.html')
    
def banner(request):
    last_activities = get_last_activity()

    if not last_activities:
        usuarios_activos = []
    else:
        usuarios = Usuario.objects.filter(id__in=[item[0] for item in last_activities])
        usuarios_activos = [
            {
                'username': usuario.username,
                'rol': usuario.rol,
                'last_activity_hour': next(item[1] for item in last_activities if item[0] == usuario.id),
                'last_activity_day': next(item[2] for item in last_activities if item[0] == usuario.id)
            }
            for usuario in usuarios
        ]
    return render(request, 'banner.html',{'usuarios_activos':usuarios_activos})

def home(request):
    datos = User.objects.all()
    messages.success(request, '¡Empleado Listado!')
    return render(request, 'inicio.html', {'Datos': datos})

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request, user)
            return redirect('/banner')
        else:
            return render(request, 'login.html', {'error': 'Credenciales inválidas'})
    else:
        return render(request, 'login.html')

def signout(request):
    logout(request)
    return redirect('login')
@user_passes_test(is_admin, login_url='/error/' )
def registro(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        rol_id = request.POST.get('rol_id')
        email =  request.POST.get('email')

        if username and rol_id:
            try:
                rol = Rol.objects.get(id=rol_id)
            except Rol.DoesNotExist:
                messages.error(request, 'El rol especificado no existe.')
                return redirect('registro')  # Corrección en la redirección

            empleado = Usuario.objects.create_user(
                username=username,
                rol=rol,
                password=password,
                email=email
            )
            messages.success(request, '¡Empleado Registrado!')
            return redirect('registro')  # Corrección en la redirección
        else:
            messages.error(request, 'Los campos "usuario" y "rol_id" no pueden estar vacíos.')
            return redirect('registro')  # Corrección en la redirección

    return render(request, 'registro.html')

def ent(request):
     return render(request, 'entrada.html')
    
def i(request):
     return render(request, 'inventario.html')

def entrada(request):
    if request.method == 'POST':
        compra = request.POST.get('compra')
        fecha = request.POST.get('fecha')
        lote = request.POST.get('lote')
        cantidad = request.POST.get('cantidad')

        # Obtener el código del producto del formulario
        codigo = request.POST.get('codigo')

        try:
            # Obtener el objeto producto relacionado con el código proporcionado
            producto_obj = producto.objects.get(codigo=codigo)
        except producto.DoesNotExist:
            # Si el producto no existe, mostrar un mensaje de error y redirigir a la página de entrada
            messages.error(request, '¡El código del producto no existe!')
            return redirect('/entrada')

        # Obtener el objeto inventario relacionado con el producto
        inventario_obj, created = inventario.objects.get_or_create(cp=producto_obj)

        # Crear la entrada
        ent = entradas.objects.create(
            compra=compra, fecha=fecha, cp=inventario_obj, lote=lote, cantidad=cantidad
        )

        # Actualizar el atributo entradas en el inventario
        inventario_obj.entradas += float(cantidad)
        inventario_obj.save()

        messages.success(request, '¡Movimiento registrado!')
        return redirect('/op')
    else:
        # Manejar la solicitud GET
        messages.warning(request, '¡Error!')
        return redirect('/ent')

#--------------------------------------maquila-------------------------------------------
@user_passes_test(is_admin_recep,login_url='/error/' )
def m(request):
    frutas = Maquila.objects.all().order_by('-id')
    emba = embalaje.objects.all()
    paletas = paleta.objects.all()
    frut = materia.objects.all()

    return render(request, 'maquila.html', {'Frutas': frutas,'emba':emba,'paletas':paletas,'frut':frut})

def enviar(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        lote = request.POST.get('lote')
        fruta_instance = fruta.objects.get(movimiento=lote)
        cajas = int(request.POST.get('cajas'))
        tipoCaja = request.POST.get('tipoCaja')
        tarima = request.POST.get('tarima')
        maquila = request.POST.get('maquila')
        cantidad = request.POST.get('cantidad')
        horaEnvio = request.POST.get('horaEnvio')
        kgNetos = request.POST.get('kgNetos')
        kgDesp= request.POST.get('kgDesp')
        hora= request.POST.get('hora')
        
        nombre = materia.objects.get(id=nombre)
        tipoCaja = embalaje.objects.get(id=tipoCaja)
        tarima = paleta.objects.get(id=tarima)

        kgNetos = Decimal(cantidad) - ((tipoCaja.peso * cajas)+tarima.peso)


        # Crear objeto en la tabla "Maquila" para registrar el envío
        maquila_registro = Maquila.objects.create(
                nombre=nombre,
                lote=fruta_instance,
                maquila=maquila,
                cantidad=cantidad,
                cajas=cajas,
                tipoCaja=tipoCaja,
                tarima=tarima,
                kgNetos=kgNetos,
                horaEnvio=horaEnvio,
                usuario=request.user,
                hora=hora
            )
        maquila_registro.save()

        messages.success(request, f'Se han enviado {cantidad} kilos y {cajas} cajas de {nombre} (lote {lote}) a la maquila {maquila}.')
    else:
        if kilos_disponibles is None:
            messages.error(request, f'No hay suficientes kilos disponibles de {nombre} (lote {lote}) en el almacén para enviar.')
        else:
            messages.error(request, f'La cantidad enviada ({cantidad_enviada}) es mayor a la cantidad disponible en el almacén ({kilos_disponibles}).')

    return redirect('m2')

def eliminarMaquila(request, maquila_id):
    try:
        maquila_obj = Maquila.objects.get(id=maquila_id)
        # Obtenemos el "movimiento" del registro a eliminar
        movimiento_eliminado = maquila_obj.id

        # Eliminamos la maquila
        maquila_obj.delete()

        # Reenumerar los IDs de todas las maquilas restantes
        maquilas_restantes = Maquila.objects.all().order_by('id')
        maquilas_restantes.update(id=F('id') - 1)

        messages.success(request, f'Se eliminó la maquila con movimiento {movimiento_eliminado} y se actualizaron los movimientos.')
    except Maquila.DoesNotExist:
        messages.error(request, 'La maquila no existe.')
    return redirect('/m2')
@user_passes_test(is_admin,login_url='/error/') 
def edicionMaquila(request, maquila_id):
    maq = Maquila.objects.get(id=maquila_id)

    if request.method == 'POST':
        lote = request.POST.get('lote')
        fruta_instance = fruta.objects.get(movimiento=lote)
        nombre = request.POST.get('nombre')
        maquila = request.POST.get('maquila')
        cajas = request.POST.get('cajas')
        tipoCaja = request.POST.get('tipoCaja')
        tarima = request.POST.get('tarima')
        cantidad_enviada = Decimal(request.POST.get('cantidad'))
        horaEnvio = request.POST.get('horaEnvio')

        maq.fruta_instance = lote
        maq.nombre = nombre
        maq.maquila = maquila
        maq.cajas = cajas
        maq.tipoCaja = tipoCaja
        maq.tarima = tarima
        maq.cantidad_enviada = cantidad_enviada
        maq.horaEnvio = horaEnvio
        
        maq.save()

        messages.success(request, '¡Movimiento Actualizado!')
        return redirect('/m2')

    return render(request, 'edicionMaquila.html', {'fruta': maq})
#------------------------------------proveedores-----------------------------------------
@user_passes_test(is_admin, login_url='/error/')
def p(request):
    frutas = proveedores.objects.all()
    return render(request, 'proveedores.html', {'Frutas': frutas})
def prov(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')

        prov = proveedores.objects.create(
            nombre=nombre
        )
        prov.save()
        messages.success(request, '¡Proveedor registrado!')
        return redirect('/p')
    else:
        # Manejar la solicitud GET
        messages.warning(request, '¡Error!')
        return redirect('/p')
def edicionProveedores(request, prov_id):
    prov =proveedores.objects.get(id=prov_id)
    if request.method == 'POST':
        nombre= request.POST.get('nombre')

        prov.nombre =nombre
        prov.save()
        messages.success(request, '!Proveedor Actualizado!')
        return redirect('/p')
    return render(request, 'edicionProveedor.html')
def eliminarProveedor(request, fruta_id):
    maquila = proveedores.objects.get(id=fruta_id)
    maquila.delete()

    messages.success(request, '¡Proveedor Eliminado!')

    return redirect('/p')
#--------------------------------Alamcen e Inventario--------------------------------------
def registrarFruta(request): 

    if request.method == 'POST':
        proveedor = request.POST.get('proveedor')
        almacen = request.POST.get('almacen')
        nombre = request.POST.get('nombre')
        unidades = Decimal(request.POST.get('unidades'))
        tipoCaja = Decimal(request.POST.get('tipoCaja'))
        tarima = Decimal(request.POST.get('tarima'))
        kg = request.POST.get('kg')
        kgNetos= request.POST.get('kgNetos')
        movimiento = request.POST.get('movimiento')
        hora = request.POST.get('hora')
        # Aquí utilizamos request.user para obtener el usuario autenticado
        usuario = request.user
        # Obtiene el ID del proveedor del formulario
        proveedor = proveedores.objects.get(id=proveedor)
        nombre = materia.objects.get(id=nombre)
        # Obtén el objeto materia correspondiente al nombre de la fruta proporcionado en el formulario
        tipoCaja = embalaje.objects.get(id=tipoCaja)
        tarima = paleta.objects.get(id=tarima)
        fecha_actual = datetime.now()
        # Formatear la fecha en el formato deseado
        fecha_formateada = fecha_actual.strftime('%Y-%m-%d')
        fecha = fecha_formateada

        
        kgNetos = Decimal(kg) - ((tipoCaja.peso * unidades) + tarima.peso)


                
        
        fecha_actual = datetime.now()
        # Formatear la fecha en el formato deseado
        fecha_formateada = fecha_actual.strftime('%Y-%m-%d')
        if almacen == 'fruta':
            # Si no hay registros en la tabla, asumimos que el valor es 0
            ultimo_movimiento = fruta.objects.order_by('-movimiento').first()
            # Si hay registros en la tabla, incrementamos el valor del movimiento en 1
            nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.movimiento + 1

            nueva_fruta = fruta.objects.create(
                movimiento=nuevo_movimiento,
                proveedor=proveedor,
                nombre_id =nombre.id,
                unidades=unidades,
                tipoCaja=tipoCaja,
                tarima=tarima,
                usuario=usuario,
                kg=kg,
                kgNetos=kgNetos,
                fecha=fecha,
                hora=hora,
                # Añade otros campos del modelo según sea necesario
            )
         

            

            return redirect('/op123')




        elif almacen == 'alfrut':
            # Si no hay registros en la tabla, asumimos que el valor es 0
            ultimo_movimiento = alfrut.objects.order_by('-movimiento').first()
            # Si hay registros en la tabla, incrementamos el valor del movimiento en 1
            nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.movimiento + 1
            nueva_alfrut = alfrut.objects.create(
                movimiento=nuevo_movimiento,
                proveedor=proveedor,
                nombre=nombre,
                unidades=unidades,
                tipoCaja=tipoCaja,
                tarima=tarima,
                usuario=usuario,
                kg=kg,
                kgNetos=kgNetos,
                fecha=fecha,
                # Añade otros campos del modelo según sea necesario
            )
            messages.success(request, '¡Movimiento registrado!')
            return redirect('/op123')

        elif almacen == 'frigoe':
            # Si no hay registros en la tabla, asumimos que el valor es 0
            ultimo_movimiento = frigoe.objects.order_by('-movimiento').first()
            # Si hay registros en la tabla, incrementamos el valor del movimiento en 1
            nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.movimiento + 1
            nueva_frigoe = frigoe.objects.create(
                movimiento=nuevo_movimiento,
                proveedor=proveedor,
                nombre=nombre,
                unidades=unidades,
                tipoCaja=tipoCaja,
                tarima=tarima,
                usuario=usuario,
                kg=kg,
                kgNetos=kgNetos,
                fecha=fecha,
                # Añade otros campos del modelo según sea necesario
            )
            messages.success(request, '¡Movimiento registrado!')
            return redirect('/op123')
        elif almacen == 'con':
            # Si no hay registros en la tabla, asumimos que el valor es 0
            ultimo_movimiento = proTer.objects.order_by('-movimiento').first()
            # Si hay registros en la tabla, incrementamos el valor del movimiento en 1
            nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.movimiento + 1
            nueva_proTer = proTer.objects.create(
                movimiento=nuevo_movimiento,
                proveedor=proveedor,
                nombre=nombre,
                unidades=unidades,
                tipoCaja=tipoCaja,
                tarima=tarima,
                usuario=usuario,
                kg=kg,
                kgNetos=kgNetos,
                fecha=fecha,
                # Añade otros campos del modelo según sea necesario
            )
            messages.success(request, '¡Movimiento registrado!')
            return redirect('/op123',{'fecha_formateada':fecha_formateada})
            
    return redirect('/op')
   
@user_passes_test(is_admin_recep, login_url='/error/')
def op(request):
    prov = proveedores.objects.all()
    ma = Maquila.objects.all().order_by('-id')
    nom = materia.objects.all()
    emba = embalaje.objects.all()
    paletas = paleta.objects.all()
    fruta_obj = fruta.objects.order_by('-movimiento').first()
    # Guardar el número de movimiento antes de eliminar la fruta
    ultimo_movimiento = fruta_obj.movimiento
    # Usar prefetch_related para obtener las relaciones de Maquila en una sola consulta
    frutas = fruta.objects.prefetch_related('maquila_set').all().order_by('-movimiento')
    if request.method == 'POST':
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')

        frutas = fruta.objects.filter(fecha__range=[fecha_inicio, fecha_fin]).order_by('-movimiento')
        return render(request, 'fruta.html',{'Frutas': frutas, 'proveedores': prov, 'maq': ma,'nombre':nom,'emba':emba,'paletas':paletas,'ultimo_movimiento':ultimo_movimiento})
    else:
        frutas = fruta.objects.all().order_by('-movimiento')
        return render(request, 'fruta.html', {'Frutas': frutas, 'proveedores': prov, 'maq': ma,'nombre':nom,'emba':emba,'paletas':paletas,'ultimo_movimiento':ultimo_movimiento})
    
    return render(request, 'fruta.html', {'Frutas': frutas, 'proveedores': prov, 'maq': ma,'nombre':nom,'emba':emba,'paletas':paletas,'ultimo_movimiento':ultimo_movimiento})

@user_passes_test(is_admin,login_url='/error/')
def edicionFruta(request, fruta_id):
    fruta_obj = fruta.objects.get(movimiento=fruta_id)
    nom = materia.objects.all()
    emba = embalaje.objects.all()
    paletas = paleta.objects.all()
    proveedores_list = proveedores.objects.all()
    if request.method == 'POST':
        proveedor = request.POST.get('proveedor')
        almacen = request.POST.get('almacen')
        nombre = request.POST.get('nombre')
        unidades = Decimal(request.POST.get('unidades'))
        tipoCaja = Decimal(request.POST.get('tipoCaja'))
        tarima = Decimal(request.POST.get('tarima'))
        kg = request.POST.get('kg')
        kgNetos= request.POST.get('kgNetos')
        movimiento = request.POST.get('movimiento')
        hora = request.POST.get('hora')
        # Aquí utilizamos request.user para obtener el usuario autenticado
        usuario = request.user
        # Obtiene el ID del proveedor del formulario
        proveedor = proveedores.objects.get(id=proveedor)
        nombre = materia.objects.get(id=nombre)
        # Obtén el objeto materia correspondiente al nombre de la fruta proporcionado en el formulario
        tipoCaja = embalaje.objects.get(id=tipoCaja)
        tarima = paleta.objects.get(id=tarima)
        fecha_actual = datetime.now()
        # Formatear la fecha en el formato deseado
        fecha_formateada = fecha_actual.strftime('%Y-%m-%d')
        fecha = fecha_formateada

        
        kgNetos = Decimal(kg) - ((tipoCaja.peso * unidades) + tarima.peso)
            
        fruta_obj.proveedor=proveedor
        fruta_obj.nombre = nombre
        fruta_obj.unidades = unidades
        fruta_obj.tipoCaja = tipoCaja
        fruta_obj.tarima=tarima
        fruta_obj.kg = kg
        fruta_obj.kgNetos = kgNetos
        fruta_obj.fecha = fecha

        fruta_obj.save()

        messages.success(request, '¡Movimiento Actualizado!')
        return redirect('/op')

    return render(request, 'edicionFruta.html', {'fruta': fruta_obj,'proveedores_list':proveedores_list,'nombre':nom,'emba':emba,'paletas':paletas})

def eliminarFruta(request, fruta_id):
    Fruta = fruta.objects.get(movimiento=fruta_id)
    Fruta.delete()

    messages.success(request, '¡Movimiento Eliminado!')

    return redirect('/op')
#------------------------------------Despate------------------------------------------------
def registrarDespate(request):
    
    if request.method == 'POST':
        lote = int(request.POST.get('lote'))
        fruta_instance = fruta.objects.get(movimiento=lote)
        nombre = request.POST.get('nombre')
        cajas = int(request.POST.get('cajas'))  # Convertir a entero
        tipoCaja = request.POST.get('tipoCaja')
        tarima = request.POST.get('tarima')
        cantidad_enviada_str = request.POST.get('kg')

        try:
            cantidad_enviada = Decimal(cantidad_enviada_str)
        except (ValueError, TypeError):
            messages.error(request, 'La cantidad debe ser un número válido y mayor que cero.')
            return redirect('op2')

        if cantidad_enviada <= 0:
            messages.error(request, 'La cantidad debe ser un número válido y mayor que cero.')
            return redirect('op2')

        horaEnvio = request.POST.get('horaEnvio')

        frutas_almacen = fruta.objects.filter(en_almacen=True, nombre=nombre, movimiento=lote)
        kilos_disponibles = frutas_almacen.aggregate(total_kilos=Sum('kg'))['total_kilos']
        cajas_disponibles = frutas_almacen.aggregate(cajas=Sum('unidades'))['cajas']
        total_kgNetos = frutas_almacen.aggregate(total_kgNetos=Sum('kgNetos'))['total_kgNetos']
        
        

        if kilos_disponibles is not None and kilos_disponibles >= cantidad_enviada:
            total_enviado_kg = 0
            total_enviado_cajas = 0

            for fruta_almacen in frutas_almacen:
                if cantidad_enviada > 0:
                    cantidad_fruta_kg = fruta_almacen.kg
                    cantidad_fruta_cajas = fruta_almacen.unidades
                    if cantidad_fruta_kg >= cantidad_enviada:
                        fruta_almacen.kg -= cantidad_enviada
                        fruta_almacen.unidades -= cajas
                        fruta_almacen.estado = 'en_maquila'
                        fruta_almacen.save()
                        total_enviado_kg += cantidad_enviada
                        total_enviado_cajas += cajas
                        cantidad_enviada = 0

                        break
                    else:
                        fruta_almacen.kg = 0
                        fruta_almacen.unidades = 0
                        fruta_almacen.en_almacen = False
                        fruta_almacen.estado = 'en_maquila'
                        fruta_almacen.save()
                        total_enviado_kg += cantidad_fruta_kg
                        total_enviado_cajas += cantidad_fruta_cajas
                        cantidad_enviada -= cantidad_fruta_kg
                        


            if tipoCaja == 'Proveedora' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 22)
            elif tipoCaja == 'Proveedora' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 26.5)
            elif tipoCaja == 'Proveedora' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 40)
            
            elif tipoCaja == 'Tijuana' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.3 + 22)
            elif tipoCaja == 'Tijuana' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.3 + 26.5)
            elif tipoCaja == 'Tijuana' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.3 + 40)

            elif tipoCaja == 'Colima' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 2.18 + 22)
            elif tipoCaja == 'Colima' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 2.18 + 26.5)
            elif tipoCaja == 'Colima' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 2.18 + 40)

            elif tipoCaja == 'Lima' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.7 + 22)
            elif tipoCaja == 'Lima' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.7 + 26.5)
            elif tipoCaja == 'Lima' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.7 + 40)

            elif tipoCaja == 'Jacona' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1 + 22)
            elif tipoCaja == 'Jacona' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1 + 26.5)
            elif tipoCaja == 'Jacona' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1 + 40)

            elif tipoCaja == 'Regilla' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 22)
            elif tipoCaja == 'Regilla' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 26.5)
            elif tipoCaja == 'Regilla' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 40)

            elif tipoCaja == 'Morelia' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.217 + 22)
            elif tipoCaja == 'Morelia' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.217 + 26.5)
            elif tipoCaja == 'Morelia' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.217 + 40)

            elif tipoCaja == 'Saltillo' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.515 + 22)
            elif tipoCaja == 'Saltillo' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.515 + 26.5)
            elif tipoCaja == 'Saltillo' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.515 + 40)

            elif tipoCaja == 'Bote' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.77 + 22)
            elif tipoCaja == 'Bote' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.77 + 26.5)
            elif tipoCaja == 'Bote' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.77 + 40)
            
            
            fruta_almacen.kgNetos -= Decimal(kgNetos)
            fruta_almacen.save()

                       

            maquila_registro = despatado.objects.create(
                nombre=nombre,
                lote=fruta_instance,
                cantidad=total_enviado_kg,
                cajas=total_enviado_cajas,
                tipoCaja=tipoCaja,
                tarima=tarima,
                kgNetos=kgNetos,
                horaEnvio=horaEnvio,
                usuario=request.user
            )
            maquila_registro.save()

            messages.success(request, f'Se han enviado {total_enviado_kg} kilos y {total_enviado_cajas} cajas de {nombre} (lote {lote}) a la maquila.')
        else:
            if kilos_disponibles is None:
                messages.error(request, f'No hay suficientes kilos disponibles de {nombre} (lote {lote}) en el almacén para enviar.')
            else:
                messages.error(request, f'La cantidad enviada ({cantidad_enviada}) es mayor a la cantidad disponible en el almacén ({kilos_disponibles}).')

    return redirect('op2')

@user_passes_test(is_admin_recep,login_url='/error/')
def op2(request):
    frutas = despatado.objects.all()
    nom = materia.objects.all()
    emba = embalaje.objects.all()
    paletas = paleta.objects.all()
    return render(request, 'despatado.html', {'Frutas': frutas})
@user_passes_test(is_admin,login_url='/error/')
def edicionDespate(request, fruta_id):
    fruta_obj = despatado.objects.get(id=fruta_id)

    if request.method == 'POST':
        lote = request.POST.get('lote')
        fruta_instance = fruta.objects.get(movimiento=lote)
        nombre =request.POST.get('nombre')
        cantidad=request.POST.get('cantidad')
        cajas=request.POST.get('cajas')
        unidades=request.POST.get('unidades')
        tipoCaja=request.POST.get('tipoCaja')
        tarima=request.POST.get('tarima')
        horaEnvio=request.POST.get('horaEnvio')

        if tipoCaja == 'Proveedora' and tarima == 'Madera':
            kgNetos = float(cantidad) - (float(cajas) * 1.68 + 22)
        elif tipoCaja == 'Proveedora' and tarima == 'Plastico':
            kgNetos = float(cantidad) - (float(cajas) * 1.68 + 26.5)
        elif tipoCaja == 'Proveedora' and tarima == 'Metal':
            kgNetos = float(cantidad) - (float(cajas) * 1.68 + 40)
            
        elif tipoCaja == 'Tijuana' and tarima == 'Madera':
            kgNetos = float(cantidad) - (float(cajas) * 1.3 + 22)
        elif tipoCaja == 'Tijuana' and tarima == 'Plastico':
            kgNetos = float(cantidad) - (float(cajas) * 1.3 + 26.5)
        elif tipoCaja == 'Tijuana' and tarima == 'Metal':
            kgNetos = float(cantidad) - (float(cajas) * 1.3 + 40)

        elif tipoCaja == 'Colima' and tarima == 'Madera':
            kgNetos = float(cantidad) - (float(cajas) * 2.18 + 22)
        elif tipoCaja == 'Colima' and tarima == 'Plastico':
            kgNetos = float(cantidad) - (float(cajas) * 2.18 + 26.5)
        elif tipoCaja == 'Colima' and tarima == 'Metal':
            kgNetos = float(cantidad) - (float(cajas) * 2.18 + 40)

        elif tipoCaja == 'Lima' and tarima == 'Madera':
            kgNetos = float(cantidad) - (float(cajas) * 0.7 + 22)
        elif tipoCaja == 'Lima' and tarima == 'Plastico':
            kgNetos = float(cantidad) - (float(cajas) * 0.7 + 26.5)
        elif tipoCaja == 'Lima' and tarima == 'Metal':
            kgNetos = float(cantidad) - (float(cajas) * 0.7 + 40)

        elif tipoCaja == 'Jacona' and tarima == 'Madera':
            kgNetos = float(cantidad) - (float(cajas) * 1 + 22)
        elif tipoCaja == 'Jacona' and tarima == 'Plastico':
            kgNetos = float(cantidad) - (float(cajas) * 1 + 26.5)
        elif tipoCaja == 'Jacona' and tarima == 'Metal':
            kgNetos = float(cantidad) - (float(cajas) * 1 + 40)

        elif tipoCaja == 'Regilla' and tarima == 'Madera':
            kgNetos = float(cantidad) - (float(cajas) * 1.68 + 22)
        elif tipoCaja == 'Regilla' and tarima == 'Plastico':
            kgNetos = float(cantidad) - (float(cajas) * 1.68 + 26.5)
        elif tipoCaja == 'Regilla' and tarima == 'Metal':
            kgNetos = float(cantidad) - (float(cajas) * 1.68 + 40)

        elif tipoCaja == 'Morelia' and tarima == 'Madera':
            kgNetos = float(cantidad) - (float(cajas) * 0.217 + 22)
        elif tipoCaja == 'Morelia' and tarima == 'Plastico':
            kgNetos = float(cantidad) - (float(cajas) * 0.217 + 26.5)
        elif tipoCaja == 'Morelia' and tarima == 'Metal':
            kgNetos = float(cantidad) - (float(cajas) * 0.217 + 40)

        elif tipoCaja == 'Saltillo' and tarima == 'Madera':
            kgNetos = float(cantidad) - (float(cajas) * 1.515 + 22)
        elif tipoCaja == 'Saltillo' and tarima == 'Plastico':
            kgNetos = float(cantidad) - (float(cajas) * 1.515 + 26.5)
        elif tipoCaja == 'Saltillo' and tarima == 'Metal':
            kgNetos = float(cantidad) - (float(cajas) * 1.515 + 40)

        elif tipoCaja == 'Bote' and tarima == 'Madera':
            kgNetos = float(cantidad) - (float(cajas) * 0.77 + 22)
        elif tipoCaja == 'Bote' and tarima == 'Plastico':
            kgNetos = float(cantidad) - (float(cajas) * 0.77 + 26.5)
        elif tipoCaja == 'Bote' and tarima == 'Metal':
            kgNetos = float(cantidad) - (float(cajas) * 0.77 + 40)
            
        fruta_obj.fruta_instance = lote
        fruta_obj.nombre = nombre
        fruta_obj.cantidad = cantidad
        fruta_obj.cajas = cajas
        fruta_obj.unidades = unidades
        fruta_obj.tipoCaja = tipoCaja
        fruta_obj.kgNetos = kgNetos
        fruta_obj.tarima = tarima
        fruta_obj.horaEnvio = horaEnvio
        
        fruta_obj.save()

        messages.success(request, '¡Movimiento Actualizado!')
        return redirect('/op2')

    return render(request, 'edicionDespate.html', {'fruta': fruta_obj})

def eliminarDespate(request, fruta_id):
    Fruta = despatado.objects.get(id=fruta_id)
    Fruta.delete()

    messages.success(request, '¡Movimiento Eliminado!')

    return redirect('/op2')
#---------------------------------------Lavado----------------------------------------------
def registrarLavado(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        lote = request.POST.get('lote')
        fruta_instance = fruta.objects.get(movimiento=lote)
        cajas = int(request.POST.get('cajas'))
        tipoCaja = request.POST.get('tipoCaja')
        tarima = request.POST.get('tarima')
        kg = request.POST.get('kg')
        horaEnvio = request.POST.get('horaEnvio')
        kgNetos = request.POST.get('kgNetos')
        kgDesp= request.POST.get('kgDesp')
        hora= request.POST.get('hora')
        nombre = materia.objects.get(id=nombre)
        tipoCaja = embalaje.objects.get(id=tipoCaja)
        tarima = paleta.objects.get(id=tarima)


        kgNetos = Decimal(kg) - ((tipoCaja.peso * cajas)+tarima.peso)

       
        maquila_registro = lavado.objects.create(
                nombre=nombre,
                lote=fruta_instance,
                kg=kg,
                cajas=cajas,
                tipoCaja=tipoCaja,
                tarima=tarima,
                kgNetos=kgNetos,
                horaEnvio=horaEnvio,
                kgDesp=kgDesp,
                usuario=request.user,
                hora=hora,
            )
        maquila_registro.save()
        messages.error(request, 'Movimiento registrado')
        return redirect('/op32')


        
    else:
        messages.error(request, 'Movimiento invalido')
        
    return redirect('op3')
@user_passes_test(is_admin_prod,login_url='/error/')
def op3(request):
    frutas = lavado.objects.all().order_by('-id')
    nom = materia.objects.all()
    emba = embalaje.objects.all()
    paletas = paleta.objects.all()
    return render(request, 'lavado.html', {'Frutas': frutas,'nombre':nom,'emba':emba,'paletas':paletas})
@user_passes_test(is_admin,login_url='/error/')
def edicionLavado(request, fruta_id):
    fruta_obj = lavado.objects.get(id=fruta_id)

    if request.method == 'POST':
        lote = request.POST.get('lote')
        fruta_instance = fruta.objects.get(id=fruta_id)
        nombre = request.POST.get('nombre')
        unidades = request.POST.get('unidades')
        tipoCaja = request.POST.get('tipoCaja')
        kg = request.POST.get('kg')
        hora_llegada = request.POST.get('hora_llegada')
        procesos = request.POST.get('procesos')
        horas_envio = request.POST.get('horas_envio')
        fecha = request.POST.get('fecha')
        hora = request.POST.get('hora')

        fruta_obj.fruta_instance = lote
        fruta_obj.nombre = nombre
        fruta_obj.unidades = unidades
        fruta_obj.tipoCaja = tipoCaja
        fruta_obj.kg = kg
        fruta_obj.hora_llegada = hora_llegada
        fruta_obj.procesos = procesos
        fruta_obj.horas_envio = horas_envio
        fruta_obj.fecha = fecha
        fruta_obj.hora = hora
        fruta_obj.save()

        messages.success(request, '¡Movimiento Actualizado!')
        return redirect('/op32')

    return render(request, 'edicionLavado.html', {'fruta': fruta_obj})

def eliminarLavado(request, fruta_id):
    ultimo_movimiento = lavado.objects.last()

    if ultimo_movimiento:
        ultimo_movimiento.delete()
        messages.success(request, '¡Movimiento Eliminado!')
    else:
        messages.success(request, '¡Movimiento no encontrado!')

    return redirect('/op42')
#--------------------------------------congelad-o----------------------------------------------
def registrarCongelado(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        lote = request.POST.get('lote')
        fruta_instance = lavado.objects.get(id=lote)
        cajas = int(request.POST.get('cajas'))
        tipoCaja = request.POST.get('tipoCaja')
        tarima = request.POST.get('tarima')
        kg = request.POST.get('kg')
        horaEnvio = request.POST.get('horaEnvio')
        kgNetos = request.POST.get('kgNetos')
        kgDesp= request.POST.get('kgDesp')
        hora= request.POST.get('hora')
        nombre = materia.objects.get(id=nombre)
        tipoCaja = embalaje.objects.get(id=tipoCaja)
        tarima = paleta.objects.get(id=tarima)


        kgNetos = Decimal(kg) - ((tipoCaja.peso * cajas)+tarima.peso)
          

        maquila_registro = congelado.objects.create(
            nombre=nombre,
            lote=fruta_instance,
            kg=kg,
            cajas=cajas,
            tipoCaja=tipoCaja,
            tarima=tarima,
            kgNetos=kgNetos,
            horaEnvio=horaEnvio,
            kgDesp=kgDesp,
            hora=hora,
            usuario=request.user,
        )
        maquila_registro.save()
        messages.error(request, f'Movimiento {lote} registrado')
        return redirect('op42')


    else:
        messages.error(request, 'Movimiento invalido')
    return redirect('/op4')

@user_passes_test(is_admin_prod,login_url='/error/')
def op4(request):
    nom = materia.objects.all()
    emba = embalaje.objects.all()
    paletas = paleta.objects.all()
    frutas = congelado.objects.all().order_by('-id')
    fruta_obj = congelado.objects.order_by('-id').first()
    # Guardar el número de movimiento antes de eliminar la fruta
    ultimo_movimiento = fruta_obj.id
    frutas2 = congelado.objects.prefetch_related('quebrado_set').all().order_by('-id')
    return render(request, 'congelado.html', {'Frutas': frutas,'frutas2':frutas2,'ultimo_movimiento':ultimo_movimiento,'nombre':nom,'emba':emba,'paletas':paletas})
@user_passes_test(is_admin,login_url='/error/')
def edicionCongelado(request, fruta_id):
    fruta_obj = congelado.objects.get(id=fruta_id)

    if request.method == 'POST':
        lote = request.POST.get('lote')
        fruta_instance = lavado.objects.get(lote=lote)
        nombre = request.POST.get('nombre')
        unidades = request.POST.get('unidades')
        tipoCaja = request.POST.get('tipoCaja')
        kg = request.POST.get('kg')
        hora_llegada = request.POST.get('hora_llegada')
        procesos = request.POST.get('procesos')
        horas_envio = request.POST.get('horas_envio')
        fecha = request.POST.get('fecha')
        hora = request.POST.get('hora')
        cliente= request.POST.get('cliente')

        fruta_obj.fruta_instance = lote
        fruta_obj.nombre = nombre
        fruta_obj.unidades = unidades
        fruta_obj.tipoCaja = tipoCaja
        fruta_obj.kg = kg
        fruta_obj.hora_llegada = hora_llegada
        fruta_obj.procesos = procesos
        fruta_obj.horas_envio = horas_envio
        fruta_obj.fecha = fecha
        fruta_obj.hora = hora
        fruta_obj.cliente = cliente
        fruta_obj.save()

        messages.success(request, '¡Movimiento Actualizado!')
        return redirect('/op42')

    return render(request, 'edicionCongelado.html', {'fruta': fruta_obj})

def eliminarCongelado(request, fruta_id):
    ultimo_movimiento = congelado.objects.last()

    if ultimo_movimiento:
        ultimo_movimiento.delete()
        messages.success(request, '¡Movimiento Eliminado!')
    else:
        messages.success(request, '¡Movimiento no encontrado!')

    return redirect('/op42')
#-----------------------------------------Quebrado----------------------------------------------
def registrarQuebrado(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        lote = request.POST.get('lote')
        fruta_instance = congelado.objects.get(id=lote)
        cajas = int(request.POST.get('cajas'))
        tipoCaja = request.POST.get('tipoCaja')
        tarima = request.POST.get('tarima')
        kg = request.POST.get('kg')
        horaEnvio = request.POST.get('horaEnvio')
        kgNetos = request.POST.get('kgNetos')
        kgDesp= request.POST.get('kgDesp')
        hora = request.POST.get('hora')
        nombre = materia.objects.get(id=nombre)
        tipoCaja = embalaje.objects.get(id=tipoCaja)
        tarima = paleta.objects.get(id=tarima)


        kgNetos = Decimal(kg) - ((tipoCaja.peso * cajas)+tarima.peso)
          

        maquila_registro = quebrado.objects.create(
            nombre=nombre,
            lote=fruta_instance,
            kg=kg,
            cajas=cajas,
            tipoCaja=tipoCaja,
            tarima=tarima,
            kgNetos=kgNetos,
            horaEnvio=horaEnvio,
            kgDesp=kgDesp,
            hora=hora,
            usuario=request.user,
        )
        maquila_registro.save()
        return redirect('op52')

    else:
        messages.error(request, 'Movimiento registrado')
    return redirect('op5')

@user_passes_test(is_admin_prod,login_url='/error/')
def op5(request):
    frutas = quebrado.objects.all().order_by('-id')
    nom = materia.objects.all()
    emba = embalaje.objects.all()
    paletas = paleta.objects.all()
    return render(request, 'quebrado.html', {'Frutas': frutas,'nombre':nom,'emba':emba,'paletas':paletas})
@user_passes_test(is_admin,login_url='/error/')
def edicionQuebrado(request, fruta_id):
    fruta_obj = quebrado.objects.get(id=fruta_id)

    if request.method == 'POST':
        movimiento = request.POST.get('movimiento')
        nombre = request.POST.get('nombre')
        unidades = request.POST.get('unidades')
        tipoCaja = request.POST.get('tipoCaja')
        kg = request.POST.get('kg')
        hora_llegada = request.POST.get('hora_llegada')
        procesos = request.POST.get('procesos')
        horas_envio = request.POST.get('horas_envio')
        fecha = request.POST.get('fecha')
        hora = request.POST.get('hora')

        fruta_obj.movimiento = movimiento
        fruta_obj.nombre = nombre
        fruta_obj.unidades = unidades
        fruta_obj.tipoCaja = tipoCaja
        fruta_obj.kg = kg
        fruta_obj.hora_llegada = hora_llegada
        fruta_obj.procesos = procesos
        fruta_obj.horas_envio = horas_envio
        fruta_obj.fecha = fecha
        fruta_obj.hora = hora
        fruta_obj.save()

        messages.success(request, '¡Movimiento Actualizado!')
        return redirect('/op52')

    return render(request, 'edicionQuebrado.html', {'fruta': fruta_obj})
 
def eliminarQuebrado(request, fruta_id):
    ultimo_movimiento = quebrado.objects.last()

    if ultimo_movimiento:
        ultimo_movimiento.delete()
        messages.success(request, '¡Movimiento Eliminado!')
    else:
        messages.success(request, '¡Movimiento no encontrado!')

    return redirect('/op42')
#-----------------------------------------concerva----------------------------------------------
def conserva(request):
    if request.method == 'POST':
        lote = request.POST.get('lote')
        tarima = request.POST.get('tarima')
        nombre = request.POST.get('nombre')
        cajas = Decimal(request.POST.get('cajas'))
        tipoCaja = request.POST.get('tipoCaja')
        fecha = request.POST.get('fecha')
        hora = request.POST.get('hora')
        #frutas_almacen = quebrado.objects.filter(nombre=nombre)
        total_kgNetos = quebrado.objects.aggregate(total_kgNetos=Sum('kgNetos'))['total_kgNetos']
        
        cliente= request.POST.get('cliente')
        nombre = materia.objects.get(id=nombre)
        tipoCaja = embalaje.objects.get(id=tipoCaja)

        kgT = (tipoCaja.peso * cajas)
        
        
        
        #if kgT <= total_kgNetos:
        cons = concervacion.objects.create(
                lote=lote, 
                tarima=tarima, 
                nombre=nombre, 
                cajas=cajas, 
                tipoCaja=tipoCaja,
                fecha=fecha, 
                cliente=cliente,
                kgT=kgT,
                hora=hora,
                aceptado=False
            )
        messages.success(request, 'Conserva registrada con éxito!!')
        cons.save()
        return redirect('con2')
            
        #else:
         #   messages.error(request, 'El valor kgNetos no fue proporcionado')
          #  return redirect('con')
    
    messages.error(request, 'El valor kgNetos no fue proporcionado')
    return redirect('con')

@user_passes_test(is_admin_recep, login_url='/error/')
def aceptar_rechazar_conserva(request, conserva_id):
    conserva = get_object_or_404(concervacion, pk=conserva_id)
    
    if request.method == 'POST':
        if 'aceptar' in request.POST:
          
                conserva.aceptado = True
                conserva.aceptador = request.user.username
                conserva.fecha_aceptacion = timezone.now()
                conserva.save()
                messages.success(request, f'Conserva aceptada por {request.user}')
        elif 'guardar_destino' in request.POST:
            nuevo_destino = request.POST.get('nuevo_destino')
            if nuevo_destino:
                conserva.destino = nuevo_destino
                conserva.save()
                messages.success(request, 'Destino actualizado exitosamente.')

        return redirect('almPT2')
    
    return render(request, 'almacenPT.html', {'Conserva': conserva})
  
@user_passes_test(is_admin_prod,login_url='/error/')
def con(request):
    frutas = concervacion.objects.all().order_by('-id')
    nom = materia.objects.all()
    emba = embalaje.objects.all()
    paletas = paleta.objects.all()
    return render(request, 'conserva.html', {'Frutas': frutas,'nombre':nom,'emba':emba,'paletas':paletas})
@user_passes_test(is_admin_prod,login_url='/error/')
def edicionConserva(request,conserva_id):
    conser = concervacion.objects.get(id=conserva_id)
    if request.method == 'POST':
        lote = request.POST.get('lote')
        tarima = request.POST.get('tarima')
        nombre = request.POST.get('nombre')
        cajas = Decimal(request.POST.get('cajas'))
        tipoCaja = request.POST.get('tipoCaja')
        fecha = request.POST.get('fecha')
        hora = request.POST.get('hora')
        #frutas_almacen = quebrado.objects.filter(nombre=nombre)
        total_kgNetos = quebrado.objects.aggregate(total_kgNetos=Sum('kgNetos'))['total_kgNetos']
        
        cliente= request.POST.get('cliente')
        nombre = materia.objects.get(id=nombre)
        tipoCaja = embalaje.objects.get(id=tipoCaja)

        kgT = (tipoCaja.peso * cajas)
        
        #if kgT <= total_kgNetos:
        conser.lote =lote
        conser.tarima=tarima
        conser.nombre=nombre
        conser.cajas=cajas
        conser.tipoCaja=tipoCaja
        conser.kgT = kgT
        conser.cliente= cliente
        conser.fecha = fecha
        conser.hora = hora

        conser.save()

        messages.success(request, '¡Movimiento Actualizado!')
        return redirect('/con2')
        #else:
         #   messages.success(request, '¡Los kilos exceden a los kilos en quebrado!')
          #  return redirect('/con')
            

    return render(request, 'edicionConserva.html', {'fruta': conser})

def eliminarConserva(request,conserva_id):
    ultimo_movimiento = concervacion.objects.last()

    if ultimo_movimiento:
        ultimo_movimiento.delete()
        messages.success(request, '¡Movimiento Eliminado!')
    else:
        messages.success(request, '¡Movimiento no encontrado!')

    return redirect('/op42')
#------------------------------------apartes--------------------------------------------------------
@user_passes_test(is_admin,login_url='/error/')
def eliminarTodo(request):
    borrar=fruta.objects.filter(kg=0)
    borrar.delete()
    messages.success(request, '¡Borrado!')
    return redirect('/op123')

def retorno(request, fruta_id):
    if request.method == 'POST':
        movimiento = request.POST.get('movimiento')
        instancia = despatado.objects.get(id=movimiento) 
        proveedor = request.POST.get('proveedor')  # Debes obtener el proveedor de alguna manera
        nombre = request.POST.get('nombre')
        unidades = int(request.POST.get('unidades'))
        tipoCaja = request.POST.get('tipoCaja')
        tarima = request.POST.get('tarima')
        fecha = request.POST.get('fecha')
        horaEnvio = request.POST.get('horaEnvio')
        cantidad_enviada = request.POST.get('kg')
        maq = 'Interna'

        try:
            cantidad_enviada = Decimal(cantidad_enviada)
        except (ValueError, TypeError):
            messages.error(request, 'La cantidad debe ser un número válido y mayor que cero.')
            return redirect('op2')

        frutas_almacen = despatado.objects.filter(en_almacen=True,  id=movimiento)
        kilos_disponibles = frutas_almacen.aggregate(total_kilos=Sum('cantidad'))['total_kilos']
        cajas_disponibles = frutas_almacen.aggregate(cajas=Sum('cajas'))['cajas']
        print(kilos_disponibles)
        print(cajas_disponibles)
        # ...
        if kilos_disponibles is not None and kilos_disponibles >= cantidad_enviada:
            total_enviado_kg = 0
            total_enviado_cajas = 0
            for fruta_almacen in frutas_almacen:
                if cantidad_enviada > 0:
                    cantidad_fruta_kg = Decimal(fruta_almacen.cantidad)
                    cantidad_fruta_cajas = fruta_almacen.cajas
                    if cantidad_fruta_kg is not None and cantidad_fruta_cajas is not None:  # Verificar si no son None
                        if cantidad_fruta_kg >= cantidad_enviada:
                            fruta_almacen.cantidad -= cantidad_enviada
                            fruta_almacen.cajas -= unidades
                            fruta_almacen.estado = 'en_almacen'
                            fruta_almacen.save()
                            total_enviado_kg += cantidad_enviada
                            total_enviado_cajas += unidades
                            cantidad_enviada = 0
                            break
                        else:
                            fruta_almacen.cantidad = 0
                            fruta_almacen.unidades = 0
                            fruta_almacen.en_almacen = False
                            fruta_almacen.estado = 'en_almacen'
                            fruta_almacen.save()
                            total_enviado_kg += cantidad_fruta_kg
                            total_enviado_cajas += cantidad_fruta_cajas
                            cantidad_enviada -= cantidad_fruta_kg


 


            if tipoCaja == 'Proveedora' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 22)
            elif tipoCaja == 'Proveedora' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 26.5)
            elif tipoCaja == 'Proveedora' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 40)
            
            elif tipoCaja == 'Tijuana' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.3 + 22)
            elif tipoCaja == 'Tijuana' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.3 + 26.5)
            elif tipoCaja == 'Tijuana' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.3 + 40)

            elif tipoCaja == 'Colima' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 2.18 + 22)
            elif tipoCaja == 'Colima' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 2.18 + 26.5)
            elif tipoCaja == 'Colima' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 2.18 + 40)

            elif tipoCaja == 'Lima' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.7 + 22)
            elif tipoCaja == 'Lima' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.7 + 26.5)
            elif tipoCaja == 'Lima' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.7 + 40)

            if tipoCaja == 'Jacona' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1 + 22)
            elif tipoCaja == 'Jacona' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1 + 26.5)
            elif tipoCaja == 'Jacona' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1 + 40)

            elif tipoCaja == 'Regilla' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 22)
            elif tipoCaja == 'Regilla' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 26.5)
            elif tipoCaja == 'Regilla' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.68 + 40)

            elif tipoCaja == 'Morelia' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.217 + 22)
            elif tipoCaja == 'Morelia' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.217 + 26.5)
            elif tipoCaja == 'Morelia' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.217 + 40)

            elif tipoCaja == 'Saltillo' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.515 + 22)
            elif tipoCaja == 'Saltillo' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.515 + 26.5)
            elif tipoCaja == 'Saltillo' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 1.515 + 40)

            elif tipoCaja == 'Bote' and tarima == 'Madera':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.77 + 22)
            elif tipoCaja == 'Bote' and tarima == 'Plastico':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.77 + 26.5)
            elif tipoCaja == 'Bote' and tarima == 'Metal':
                kgNetos = float(total_enviado_kg) - (float(total_enviado_cajas) * 0.77 + 40)
            
            print(total_enviado_kg)
            print(total_enviado_cajas)
            
             # Si no hay registros en la tabla, asumimos que el valor es 0
            ultimo_movimiento = fruta.objects.order_by('-movimiento').first()

            # Si hay registros en la tabla, incrementamos el valor del movimiento en 1
            nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.movimiento + 1


            proveedor_obj= proveedores.objects.get(id=8)
            # Crear objeto en la tabla "Maquila" para registrar el envío
            maquila_registro = fruta.objects.create(
                movimiento=nuevo_movimiento,
                nombre=nombre,
                proveedor=proveedor_obj,
                kg=total_enviado_kg,
                unidades=total_enviado_cajas,
                tipoCaja=tipoCaja,
                tarima=tarima,
                kgNetos=kgNetos,
                fecha=fecha,
                maq=maq,
                usuario=request.user
            )
            maquila_registro.save()

            messages.success(request, f'Han ingresado {total_enviado_kg} kilos y {total_enviado_cajas} cajas de {nombre} al almacén.')
        
        else:
            if kilos_disponibles is None:
                messages.error(request, f'No hay suficientes kilos disponibles de {nombre}  en la maquila.')
            else:
                messages.error(request, f'La cantidad de entrada ({cantidad_enviada}) es mayor a la cantidad disponible en la maquila ({kilos_disponibles}).')

    return redirect('ent')

def retornoEx(request, fruta_id):
    if request.method == 'POST':
        proveedor = request.POST.get('proveedor')
        almacen = request.POST.get('almacen')
        nombre = request.POST.get('nombre')
        movimiento = request.POST.get('movimiento')
        instancia = Maquila.objects.get(id=movimiento) 
        unidades = request.POST.get('unidades')
        tipoCaja = request.POST.get('tipoCaja')
        tarima = request.POST.get('tarima')
        kg = request.POST.get('kg')
        kgNetos= request.POST.get('kgNetos')
        fecha = request.POST.get('fecha')
        fecha1 = request.POST.get('fecha1')
        movimiento = request.POST.get('movimiento')
        maquila_origen = request.POST.get('maquila_origen')
        hora = request.POST.get('hora')
       
        # Aquí utilizamos request.user para obtener el usuario autenticado
        usuario = request.user
        # Obtiene el ID del proveedor del formulario
        
        maquila_origen = movimiento
        tipoCaja = embalaje.objects.get(id=tipoCaja)
        tarima = paleta.objects.get(id=tarima)
      
        kgNetos = Decimal(kg) - ((tipoCaja.peso * Decimal(unidades))+tarima.peso)
        
            # Si no hay registros en la tabla, asumimos que el valor es 0
        ultimo_movimiento = fruta.objects.order_by('-movimiento').first()

            # Si hay registros en la tabla, incrementamos el valor del movimiento en 1
        nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.movimiento + 1


        proveedor_obj= proveedores.objects.get(id=8)
        nombre = materia.objects.get(id=nombre)
        frutas_almacen = Maquila.objects.filter(en_almacen=True, nombre=nombre, id=movimiento)

            # Obtener un objeto Maquila específico del QuerySet (por ejemplo, el primer objeto)
        maquila_obj = frutas_almacen.first()

            # Verificar si se encontró un objeto Maquila en el QuerySet
        if maquila_obj:
                # Obtener el valor del atributo 'maquila' del objeto Maquila
            maq = maquila_obj.maquila
        else:
                # Manejar el caso si no se encontró ningún objeto Maquila en el QuerySet
            maq = None



            # Crear objeto en la tabla "Maquila" para registrar el envío
        maquila_registro = fruta.objects.create(
                movimiento=nuevo_movimiento,
                nombre=nombre,
                proveedor= proveedor_obj,
                kg=kg,
                unidades=unidades,
                tipoCaja=tipoCaja,
                tarima=tarima,
                kgNetos=kgNetos,
                fecha=fecha,
                usuario=request.user,
                hora=hora,
                maq = maq,
                maquila_origen = maquila_origen
            )
        maquila_registro.save()

        messages.success(request, f'Han ingresado {kg} kilos y {unidades} cajas de {nombre} al almacén.')
        
    else:
        if kilos_disponibles is None:
            messages.error(request, f'No hay suficientes kilos disponibles de {nombre}  en la maquila.')
        else:
            messages.error(request, f'La cantidad de entrada ({cantidad_enviada}) es mayor a la cantidad disponible en la maquila ({kilos_disponibles}).')

    return redirect('entEx')

def ent(request):
    fruta_id = despatado.id  # Reemplaza esto con el valor real que necesitas
    return render(request, 'entrada.html', {'fruta_id': fruta_id})

def entExt(request):
    fruta_id = despatado.id  # Reemplaza esto con el valor real que necesitas
    # Obtener los IDs de las maquilas que están en maquila_origen en la tabla Fruta
    maquilas_en_maq_origen = fruta.objects.exclude(maquila_origen=None).values_list('maquila_origen', flat=True)

    # Excluir las maquilas que ya están en maquila_origen
    frutas = Maquila.objects.exclude(id__in=maquilas_en_maq_origen).order_by('-id')
    frut = materia.objects.all()
    emba = embalaje.objects.all()
    paletas = paleta.objects.all()
    return render(request, 'entradaEx.html', {'fruta_id': fruta_id,'Frutas':frutas,'frut':frut,'emba':emba,'paletas':paletas})

@user_passes_test(is_admin_prod,login_url='/error/')
def verAl(request): 
    nombre_filtro = request.GET.get('nombre_filtro', '')  # Obtener el valor del parámetro GET
    frutas = fruta.objects.all().order_by('-movimiento')
    
    if nombre_filtro:
        frutas_filtradas = frutas.filter(nombre=nombre_filtro)
        total_cajas = frutas_filtradas.aggregate(total_cajas=Sum('unidades'))['total_cajas']
        total_kgNetos = frutas_filtradas.aggregate(total_kgNetos=Sum('kgNetos'))['total_kgNetos']
    else:
        total_cajas = 0
        total_kgNetos = 0

    context = {
        'Frutas': frutas,
        'total_cajas': total_cajas,
        'total_kgNetos': total_kgNetos,
        'nombre_filtro': nombre_filtro,  # Pasar el valor del filtro a la plantilla
    }

    return render(request, 'almacen.html', context)

#----------------------------------excel papeletas------------------------------------------------------
def descargar_excel(request, fruta_id):
    # Filtrar los registros de fruta para el movimiento específico
    frutas = fruta.objects.filter(movimiento=fruta_id)

    if not frutas.exists():
        # Manejar el caso en el que no hay registros para el movimiento dado
        # Puedes redirigir a una página de error o mostrar un mensaje apropiado
        return HttpResponse('No hay registros para el movimiento {}'.format(fruta_id))

    # Cargar el archivo Excel preexistente
    workbook = openpyxl.load_workbook('C:/Users/Facturacion/Desktop/PRUEBAS/Aplicaciones/Trabajo/papeleta.xlsx')
    sheet = workbook.active

    # Obtener la fecha actual
    fecha_actual = datetime.now()

    # Formatear la fecha en el formato deseado
    fecha_formateada = fecha_actual.strftime('%Y-%m-%d %H:%M:%S')
    # Asignar los datos a celdas específicas (usando el primer registro)
    fruta_obj = frutas.first()  # Puedes ajustar esto según tus necesidades
    datos = f'{fruta_obj.proveedor.nombre},{fruta_obj.maq}'

    sheet['C9'] = datos
    sheet['C10'] = fruta_obj.movimiento
    sheet['C11'] = fruta_obj.kgNetos
    sheet['C12'] = fruta_obj.unidades
    sheet['C14'] = fruta_obj.usuario.username
    sheet['H9'] = fecha_formateada
    sheet['H10'] = fruta_obj.tarima
    sheet['H11'] = fruta_obj.tipoCaja
    sheet['H12'] = fruta_obj.kg
    sheet['H13'] = fruta_obj.nombre
    sheet['C27'] = datos
    sheet['C28'] = fruta_obj.movimiento
    sheet['C29'] = fruta_obj.kgNetos
    sheet['C30'] = fruta_obj.unidades
    sheet['C32'] = fruta_obj.usuario.username
    sheet['H31'] = fruta_obj.nombre
    sheet['H27'] = fecha_formateada
    sheet['H28'] = fruta_obj.tarima
    sheet['H29'] = fruta_obj.tipoCaja
    sheet['H30'] = fruta_obj.kg

    # Establecer propiedades de protección para evitar la edición del archivo
    for row in sheet.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    # Proteger la hoja
    sheet.protection.sheet = True

    # Guardar el archivo Excel modificado y enviarlo como respuesta
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=papeleta_tarima_{}.xlsx'.format(fruta_id)

    return response
def descargar_excel2(request, fruta_id):
    # Filtrar los registros de fruta para el movimiento específico
    frutas = despatado.objects.filter(id=fruta_id)

    if not frutas.exists():
        # Manejar el caso en el que no hay registros para el movimiento dado
        # Puedes redirigir a una página de error o mostrar un mensaje apropiado
        return HttpResponse('No hay registros para el movimiento {}'.format(fruta_id))

    # Cargar el archivo Excel preexistente
    workbook = openpyxl.load_workbook('C:/Users/Facturacion/Desktop/PRUEBAS/Aplicaciones/Trabajo/papeletaPT')
    sheet = workbook.active

    # Obtener la fecha actual
    fecha_actual = datetime.now()

    # Formatear la fecha en el formato deseado
    fecha_formateada = fecha_actual.strftime('%Y-%m-%d %H:%M:%S')
    # Asignar los datos a celdas específicas (usando el primer registro)
    fruta_obj = frutas.first()  # Puedes ajustar esto según tus necesidades

    sheet['C10'] = fruta_obj.id
    sheet['C11'] = fruta_obj.kgNetos
    sheet['C12'] = fruta_obj.cajas
    sheet['C14'] = fruta_obj.usuario.username
    sheet['H9'] = fecha_formateada
    sheet['H10'] = fruta_obj.tarima
    sheet['H11'] = fruta_obj.tipoCaja
    sheet['H12'] = fruta_obj.cantidad
    sheet['C28'] = fruta_obj.id
    sheet['C29'] = fruta_obj.kgNetos
    sheet['C30'] = fruta_obj.cajas
    sheet['C32'] = fruta_obj.usuario.username
    sheet['H27'] = fecha_formateada
    sheet['H28'] = fruta_obj.tarima
    sheet['H29'] = fruta_obj.tipoCaja
    sheet['H30'] = fruta_obj.cantidad

    # Establecer propiedades de protección para evitar la edición del archivo
    for row in sheet.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    # Proteger la hoja
    sheet.protection.sheet = True

    # Guardar el archivo Excel modificado y enviarlo como respuesta
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=papeleta_tarima_{}.xlsx'.format(fruta_id)

    return response

def descargar_excelPT(request, fruta_id):
    # Filtrar los registros de fruta para el movimiento específico
    frutas = congelado.objects.filter(id=fruta_id)


    if not frutas.exists():
        # Manejar el caso en el que no hay registros para el movimiento dado
        # Puedes redirigir a una página de error o mostrar un mensaje apropiado
        return HttpResponse('No hay registros para el movimiento {}'.format(fruta_id))

    # Cargar el archivo Excel preexistente
    workbook = openpyxl.load_workbook('C:/Users/Facturacion/Desktop/PRUEBAS/Aplicaciones/Trabajo/papeletaPT.xlsx')
    sheet = workbook.active

    # Obtener la fecha actual
    fecha_actual = datetime.now()

    # Formatear la fecha en el formato deseado
    fecha_formateada = fecha_actual.strftime('%Y-%m-%d %H:%M:%S')
    # Asignar los datos a celdas específicas (usando el primer registro)
    fruta_obj = frutas.first()
    sheet['B9'] = fruta_obj.nombre
    sheet['B10'] = fruta_obj.cajas
    sheet['B11'] = fruta_obj.lote_id
    sheet['B12'] = fruta_obj.cliente
    sheet['K8'] = fecha_formateada
    sheet['L12'] = fruta_obj.id
    sheet['I10'] = fruta_obj.kg
    sheet['I11'] = fruta_obj.kgNetos
    sheet['H9'] = fruta_obj.tipoCaja
    sheet['A14'] = request.user.username


    sheet['B27'] = fruta_obj.nombre
    sheet['B28'] = fruta_obj.cajas
    sheet['B29'] = fruta_obj.lote_id
    sheet['B30'] = fruta_obj.cliente
    sheet['K8'] = fecha_formateada
    sheet['L30'] = fruta_obj.id
    sheet['I28'] = fruta_obj.kg
    sheet['I29'] = fruta_obj.kgNetos
    sheet['H27'] = fruta_obj.tipoCaja
    sheet['A32'] = request.user.username



    # Establecer propiedades de protección para evitar la edición del archivo
    for row in sheet.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    # Proteger la hoja
    sheet.protection.sheet = True

    # Guardar el archivo Excel modificado y enviarlo como respuesta
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=papeleta_tarima_{}.xlsx'.format(fruta_id)

    return response

def descargar_excelPT2(request, fruta_id):
    # Filtrar los registros de fruta para el movimiento específico
    frutas = concervacion.objects.filter(id=fruta_id)

    if not frutas.exists():
        # Manejar el caso en el que no hay registros para el movimiento dado
        # Puedes redirigir a una página de error o mostrar un mensaje apropiado
        return HttpResponse('No hay registros para el movimiento {}'.format(fruta_id))

    # Cargar el archivo Excel preexistente
    workbook = openpyxl.load_workbook('C:/Users/Facturacion/Desktop/PRUEBAS/Aplicaciones/Trabajo/papeletaPT.xlsx')
    sheet = workbook.active

    # Obtener la fecha actual
    fecha_actual = datetime.now()

    # Formatear la fecha en el formato deseado
    fecha_formateada = fecha_actual.strftime('%Y-%m-%d %H:%M:%S')
    # Asignar los datos a celdas específicas (usando el primer registro)
    fruta_obj = frutas.first()  # Puedes ajustar esto según tus necesidades

    print("Valor de fruta_obj.kgT:", fruta_obj.kgT) 


    sheet['B9'] = fruta_obj.nombre
    sheet['B10'] = fruta_obj.cajas
    sheet['B11'] = fruta_obj.lote
    sheet['B12'] = fruta_obj.cliente
    sheet['K8'] = fecha_formateada
    sheet['L12'] = fruta_obj.tarima
    sheet['I11'] = fruta_obj.kgT
    sheet['I9'] = fruta_obj.tipoCaja

    sheet['B27'] = fruta_obj.nombre
    sheet['B28'] = fruta_obj.cajas
    sheet['B29'] = fruta_obj.lote
    sheet['B30'] = fruta_obj.cliente
    sheet['K8'] = fecha_formateada
    sheet['L30'] = fruta_obj.tarima
    sheet['I27'] = fruta_obj.tipoCaja
    sheet['I29'] = fruta_obj.kgT

    # Establecer propiedades de protección para evitar la edición del archivo
    for row in sheet.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

        # Proteger la hoja
    sheet.protection.sheet = True

    # Guardar el archivo Excel modificado y enviarlo como respuesta
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=papeleta_tarima_{}.xlsx'.format(fruta_id)

    return response

#----------------------------------------------------almacenes-------------------------------------------
@user_passes_test(is_admin_recep,login_url='/error/')
def almAlf(request):
    frutas = alfrut.objects.all()
    return render(request, 'almacenAlf.html', {'Frutas': frutas})
@user_passes_test(is_admin_recep,login_url='/error/')
def almPT(request):
    nombre_filtro = request.GET.get('nombre_filtro', '')  # Obtener el valor del parámetro GET
    status = request.GET.get('status', '')  # Obtener el valor del parámetro GET
    frutas = concervacion.objects.all().order_by('-id')
    
    if nombre_filtro:
        frutas_filtradas = frutas.filter(nombre=nombre_filtro,destino=status)
        total_cajas = frutas_filtradas.aggregate(total_cajas=Sum('cajas'))['total_cajas']
        total_kgNetos = frutas_filtradas.aggregate(total_kgNetos=Sum('kgT'))['total_kgNetos']
    else:
        total_cajas = 0
        total_kgNetos = 0

    context = {
        'Frutas': frutas,
        'total_cajas': total_cajas,
        'total_kgNetos': total_kgNetos,
        'nombre_filtro': nombre_filtro,  # Pasar el valor del filtro a la plantilla
    }

    return render(request, 'almacenPT.html', context)
@user_passes_test(is_admin_recep,login_url='/error/')
def almFri(request):
    frutas = frigoe.objects.all() 
    return render(request, 'almacenFri.html', {'Frutas': frutas})
@user_passes_test(is_admin_recep,login_url='/error/')
def almAli(request):
    frutas = fruta.objects.all().order_by('-movimiento')
    return render(request, 'almacenAli.html', {'Frutas': frutas})
@user_passes_test(is_admin_recep,login_url='/error/')
def verCon(request):
    frutas = congelado.objects.all()
    return render(request, 'congeladoAlm.html', {'Frutas': frutas})
@user_passes_test(is_admin_recep,login_url='/error/')
def tras(request):
    return render(request, 'traspasos.html')
@user_passes_test(is_admin_recep,login_url='/error/')
def traspasarMovimiento(request):
    if request.method == 'POST':
        # Obtener el ID del movimiento a transferir
        movimiento_id = request.POST.get('movimiento_id')
        # Obtener el almacén de origen y el almacén de destino
        almacen_origen = request.POST.get('almacen_origen')
        almacen_destino = request.POST.get('almacen_destino')

        

        if almacen_origen == almacen_destino:
            messages.error(request, 'El almacén de origen y destino no pueden ser iguales.')
            return redirect('/tras2')

        try:
            # Recuperar el movimiento a partir del ID en la tabla de origen
            if almacen_origen == 'alfrut':
                movimiento_id = alfrut.objects.get(pk=movimiento_id)
            elif almacen_origen == 'frigoe':
                movimiento_id = frigoe.objects.get(pk=movimiento_id)
            elif almacen_origen == 'con':
                movimiento_id = proTer.objects.get(pk=movimiento_id)
            elif almacen_origen == 'fruta':
                movimiento_id = fruta.objects.get(pk=movimiento_id)
        except alfrut.DoesNotExist:
            messages.error(request, 'Ese movimiento no existe en el almacén origen ' + almacen_origen)
            return redirect('/tras2')

        # Crear un nuevo registro en la tabla de destino
        if almacen_destino == 'fruta':
            # Si no hay registros en la tabla, asumimos que el valor es 0
            ultimo_movimiento = fruta.objects.order_by('-movimiento').first()
            # Si hay registros en la tabla, incrementamos el valor del movimiento en 1
            nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.movimiento + 1
            nuevo_movimiento1 = fruta.objects.create(
                movimiento=nuevo_movimiento,
                proveedor=movimiento_id.proveedor,
                nombre=movimiento_id.nombre,
                unidades=movimiento_id.unidades,
                tipoCaja=movimiento_id.tipoCaja,
                tarima=movimiento_id.tarima,
                usuario=movimiento_id.usuario,
                kg=movimiento_id.kg,
                kgNetos=movimiento_id.kgNetos,
                fecha=movimiento_id.fecha,
                # Añade otros campos del modelo según sea necesario
            )
        elif almacen_destino == 'alfrut':
            # Si no hay registros en la tabla, asumimos que el valor es 0
            ultimo_movimiento = alfrut.objects.order_by('-movimiento').first()
            # Si hay registros en la tabla, incrementamos el valor del movimiento en 1
            nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.movimiento + 1
            nuevo_movimiento1 = alfrut.objects.create(
                movimiento=nuevo_movimiento,
                proveedor=movimiento_id.proveedor,
                nombre=movimiento_id.nombre,
                unidades=movimiento_id.unidades,
                tipoCaja=movimiento_id.tipoCaja,
                tarima=movimiento_id.tarima,
                usuario=movimiento_id.usuario,
                kg=movimiento_id.kg,
                kgNetos=movimiento_id.kgNetos,
                fecha=movimiento_id.fecha,
                # Añade otros campos del modelo según sea necesario
            )
        elif almacen_destino == 'con':
            # Si no hay registros en la tabla, asumimos que el valor es 0
            ultimo_movimiento = proTer.objects.order_by('-movimiento').first()
            # Si hay registros en la tabla, incrementamos el valor del movimiento en 1
            nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.movimiento + 1
            nuevo_movimiento1 = proTer.objects.create(
                movimiento=nuevo_movimiento,
                proveedor=movimiento_id.proveedor,
                nombre=movimiento_id.nombre,
                unidades=movimiento_id.unidades,
                tipoCaja=movimiento_id.tipoCaja,
                tarima=movimiento_id.tarima,
                usuario=movimiento_id.usuario,
                kg=movimiento_id.kg,
                kgNetos=movimiento_id.kgNetos,
                fecha=movimiento_id.fecha,
                # Añade otros campos del modelo según sea necesario
            )
        elif almacen_destino == 'frigoe':
            # Si no hay registros en la tabla, asumimos que el valor es 0
            ultimo_movimiento = frigoe.objects.order_by('-movimiento').first()
            # Si hay registros en la tabla, incrementamos el valor del movimiento en 1
            nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.movimiento + 1
            nuevo_movimiento1 = frigoe.objects.create(
                movimiento=nuevo_movimiento,
                proveedor=movimiento_id.proveedor,
                nombre=movimiento_id.nombre,
                unidades=movimiento_id.unidades,
                tipoCaja=movimiento_id.tipoCaja,
                tarima=movimiento_id.tarima,
                usuario=movimiento_id.usuario,
                kg=movimiento_id.kg,
                kgNetos=movimiento_id.kgNetos,
                fecha=movimiento_id.fecha,
                # Añade otros campos del modelo según sea necesario
            )
        elif almacen_destino == 'ventas':
            # Si no hay registros en la tabla, asumimos que el valor es 0
            ultimo_movimiento = alfrut.objects.order_by('-movimiento').first()
            # Si hay registros en la tabla, incrementamos el valor del movimiento en 1
            nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.movimiento + 1
            nuevo_movimiento1 = ventas.objects.create(
                movimiento=nuevo_movimiento,
                proveedor=movimiento_id.proveedor,
                nombre=movimiento_id.nombre,
                unidades=movimiento_id.unidades,
                tipoCaja=movimiento_id.tipoCaja,
                tarima=movimiento_id.tarima,
                usuario=movimiento_id.usuario,
                kg=movimiento_id.kg,
                kgNetos=movimiento_id.kgNetos,
                fecha=movimiento_id.fecha,
                # Añade otros campos del modelo según sea necesario
            )

        # Eliminar el registro original del movimiento en la tabla de origen
        movimiento_id.delete()

        messages.success(request, '¡Movimiento transferido!')
        return redirect('/tras2')
    else:
        # Manejar la solicitud GET
        return redirect('/tras2')
@user_passes_test(is_admin,login_url='/error/')
def vent(request):
    frutas = ventas.objects.all()
    return render(request, 'ventas.html', {'Frutas': frutas})

#-----------------------------------------------ordenes de de produccion-------------------------------------
@user_passes_test(is_admin, login_url='/error/' )
def orden(request):
    if request.method == 'POST':
        cantidad = request.POST.get('cantidad')
        clave = request.POST.get('clave')
        producto = request.POST.get('producto')
        fecha = request.POST.get('fecha')
        po = request.POST.get('po')
        usuario = request.user

        o = ordenes.objects.create(
            cantidad=cantidad,clave=clave,producto=producto,fecha=fecha,po=po,usuario=usuario
        )
        o.save()
        messages.success(request, 'Orden registrada con éxito!!')
        return redirect('orden1')
    else:
        messages.error(request, 'La orden no se pudo crear')
        return redirect('orden1')
@user_passes_test(is_admin, login_url='/error/' )
def orden1(request):
    ordd= ordenes.objects.all()
    return render(request, 'ordenes.html', {'Frutas': ordd})
@user_passes_test(is_admin, login_url='/error/' )
def edicionOrden(request,fruta_id):
    conser = ordenes.objects.get(id=fruta_id)
    if request.method == 'POST':
        cantidad = request.POST.get('cantidad')
        clave = request.POST.get('clave')
        producto = request.POST.get('producto')
        fecha = request.POST.get('fecha')
        po = request.POST.get('po')

        
        conser.cantidad =cantidad
        conser.clave=clave
        conser.producto=producto
        conser.fecha=fecha
        conser.po=po
        
        conser.save()

        messages.success(request, '¡Movimiento Actualizado!')
        return redirect('/orden1')
      

    return render(request, 'edicionOrden.html', {'fruta': conser})
def eliminarOrden(request,fruta_id):
    Fruta = ordenes.objects.get(id=fruta_id)
    Fruta.delete()

    messages.success(request, '¡Movimiento Eliminado!')

    return redirect('/orden1')
def aceptar_orden(request, fruta_id):
    conserva = get_object_or_404(ordenes, pk=fruta_id)
    
    if request.method == 'POST':
        if 'completa' in request.POST:
            print("Formulario enviado con éxito")  # Agrega este registro de depuración
            
            conserva.aceptado = True
            conserva.aceptador = request.user.username
            conserva.fecha_aceptacion = timezone.now()
            conserva.save()
            messages.success(request, f'Orden terminada!!')
            return redirect('or12')
        else:
            messages.error(request, 'La orden no se acepto')
            return redirect('or12')

    
    return render(request, 'orden2.html', {'Conserva': conserva})
@user_passes_test(is_admin_prod, login_url='/error/' )
def or1(request):
    ordd= ordenes.objects.all()
    return render(request, 'orden2.html', {'Frutas': ordd})

#---------------------------------------------------USUARIOS---------------------------------------------------
def us(request):
    user = request.user
        # Obtiene la contraseña del usuario
    user_password = request.user.password

    # Máscara la contraseña (cambia los caracteres por asteriscos)
    masked_password = '*' * len(default_if_none(user_password, ''))

    # Renderiza la plantilla con la contraseña enmascarada
    return render(request, 'user.html', {'masked_password': masked_password})
    return render(request, 'user.html', {'user': user})
        
def u(request):
    user = request.user
    return render(request, 'user.html', {'user': user})
@login_required
def guardar_cambios(request):
    if request.method == 'POST':
        # Obtén los datos del formulario
        new_username = request.POST['username']
        new_password = request.POST['password']
        new_nombres = request.POST['nombres']
        new_apellidos = request.POST['apellidos']
        fotos_perfil = request.FILES.get('fotos_perfil')
        try:
            foto_perfil = request.FILES['fotos_perfil']  # Accede al archivo de imagen cargado
        except MultiValueDictKeyError:
            foto_perfil = None  # Si no se proporciona ninguna imagen, foto_perfil será None
        
        # Actualiza los datos del usuario
        user = request.user
        user.username = new_username
        user.set_password(new_password)  # Debes hashear la nueva contraseña
        user.nombres = new_nombres
        user.apellidos = new_apellidos
        user.foto=fotos_perfil
        user.save()

        
        return redirect('u')

        # Agrega un mensaje de éxito
        messages.success(request, 'Cambios guardados exitosamente.')

    return redirect('/u')

@user_passes_test(is_admin, login_url='/error/' )
def compra(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        cantidad = request.POST.get('cantidad')
        fecha = request.POST.get('fecha')
        Usuario = request.user

        en = compras.objects.create(
            nombre=nombre,cantidad=cantidad,fecha=fecha
        )
        en.save()
        messages.success(request,'Compra guardada!!')
        return redirect('com')

    return redirect('/com' )
@user_passes_test(is_admin, login_url='/error/' )
def com(request):
    en = compras.objects.all()
    return render(request, 'compras.html', {'Frutas': en})
@user_passes_test(is_admin, login_url='/error/' )
def eliminarCompra(request,fruta_id):
    Fruta = compras.objects.get(id=fruta_id)
    Fruta.delete()

    messages.success(request, '¡Movimiento Eliminado!')

def rend(request):
    return render(request, 'rendimientos.html')

#-------------------------------------------------cajas------------------------------------------
@login_required
def registrar_salida(request):
    if request.method == 'POST':
        proveedor_id = request.POST.get('proveedor')
        unidades = int(request.POST.get('unidades'))
        usuario = request.user

        proveedor = proveedores.objects.get(id=proveedor_id)
        recalculate_saldos_cajas()
        
        if proveedor.saldo_cajas is None:
            proveedor.saldo_cajas = 0

        if proveedor.saldo_cajas >= unidades:
            proveedor.saldo_cajas -= unidades
            proveedor.save()

            salida = SalidaStock.objects.create(
                proveedor=proveedor,
                unidades=unidades,
                fecha=timezone.now(),
                usuario=usuario
            )

            return redirect('registrar_salida')  # Redirige a la página que desees después de registrar la salida
        else:
            # Manejar el caso donde no hay suficientes unidades en el saldo de cajas
            error_message = "No hay suficientes unidades en el saldo de cajas"
            return render(request, 'registrar_salida.html', {'proveedores': proveedores.objects.all(), 'error_message': error_message})

    return render(request, 'registrar_salida.html', {'proveedores': proveedores.objects.all()})


def recalculate_saldos_cajas():
    # Restablece el saldo de cajas de todos los proveedores a cero
    proveedores_list = proveedores.objects.all()
    for proveedor in proveedores_list:
        proveedor.saldo_cajas = 0
        proveedor.save()

    # Suma las unidades de todas las entradas de fruta
    frutas = fruta.objects.all()
    for entrada in frutas:
        proveedor = entrada.proveedor
        proveedor.saldo_cajas += entrada.unidades
        proveedor.save()

    # Resta las unidades de todas las salidas de stock
    salidas = SalidaStock.objects.all()
    for salida in salidas:
        proveedor = salida.proveedor
        proveedor.saldo_cajas -= salida.unidades
        proveedor.save()

    return HttpResponse("Saldos recalculados exitosamente")


#--------------------------------------------------conglado a carfo de recepcion----------------------------------
@user_passes_test(is_admin_prod, login_url='/error/' )
def coPro(request):
    frutas = congeladoP.objects.all().order_by('-Con')
    nombre = materia.objects.all()
    return render(request, 'congeladoP.html', {'Frutas': frutas,'nombre':nombre})

@user_passes_test(is_admin_prod, login_url='/error/' )
def congeladoPro(request):
    if request.method == 'POST':
        pro = request.POST.get('pro')
        nombre = request.POST.get('nombre')
        cajas = int(request.POST.get('cajas'))
        tipoCaja = request.POST.get('tipoCaja')
        tarima = request.POST.get('tarima')
        kg = request.POST.get('kg')
        status = request.POST.get('status')
        horaEnvio = request.POST.get('horaEnvio')
        kgNetos = request.POST.get('kgNetos')
        kgDesp= request.POST.get('kgDesp')
        cliente= request.POST.get('cliente')
        if 'sin_cliente' in request.POST:
            cliente = 'Sin Cliente'
        
        nombre = materia.objects.get(id = nombre)

        if tipoCaja == 'Proveedora' and tarima == 'Madera':
            kgNetos = float(kg) - (float(cajas) * 1.68 + 22)
        elif tipoCaja == 'Proveedora' and tarima == 'Plastico':
            kgNetos = float(kg) - (float(cajas) * 1.68 + 26.5)
        elif tipoCaja == 'Proveedora' and tarima == 'Metal':
            kgNetos = float(kg) - (float(cajas) * 1.68 + 40)
            
        elif tipoCaja == 'Tijuana' and tarima == 'Madera':
            kgNetos = float(kg) - (float(cajas) * 1.3 + 22)
        elif tipoCaja == 'Tijuana' and tarima == 'Plastico':
            kgNetos = float(kg) - (float(cajas) * 1.3 + 26.5)
        elif tipoCaja == 'Tijuana' and tarima == 'Metal':
            kgNetos = float(kg) - (float(cajas) * 1.3 + 40)

        elif tipoCaja == 'Colima' and tarima == 'Madera':
            kgNetos = float(kg) - (float(cajas) * 2.18 + 22)
        elif tipoCaja == 'Colima' and tarima == 'Plastico':
            kgNetos = float(kg) - (float(cajas) * 2.18 + 26.5)
        elif tipoCaja == 'Colima' and tarima == 'Metal':
            kgNetos = float(kg) - (float(cajas) * 2.18 + 40)

        elif tipoCaja == 'Lima' and tarima == 'Madera':
            kgNetos = float(kg) - (float(cajas) * 0.7 + 22)
        elif tipoCaja == 'Lima' and tarima == 'Plastico':
            kgNetos = float(kg) - (float(cajas) * 0.7 + 26.5)
        elif tipoCaja == 'Lima' and tarima == 'Metal':
            kgNetos = float(kg) - (float(cajas) * 0.7 + 40)

        elif tipoCaja == 'Jacona' and tarima == 'Madera':
            kgNetos = float(kg) - (float(cajas) * 1 + 22)
        elif tipoCaja == 'Jacona' and tarima == 'Plastico':
            kgNetos = float(kg) - (float(cajas) * 1 + 26.5)
        elif tipoCaja == 'Jacona' and tarima == 'Metal':
            kgNetos = float(kg) - (float(cajas) * 1 + 40)

        elif tipoCaja == 'Regilla' and tarima == 'Madera':
            kgNetos = float(kg) - (float(cajas) * 1.68 + 22)
        elif tipoCaja == 'Regilla' and tarima == 'Plastico':
            kgNetos = float(kg) - (float(cajas) * 1.68 + 26.5)
        elif tipoCaja == 'Regilla' and tarima == 'Metal':
            kgNetos = float(kg) - (float(cajas) * 1.68 + 40)

        elif tipoCaja == 'Morelia' and tarima == 'Madera':
            kgNetos = float(kg) - (float(cajas) * 0.217 + 22)
        elif tipoCaja == 'Morelia' and tarima == 'Plastico':
            kgNetos = float(kg) - (float(cajas) * 0.217 + 26.5)
        elif tipoCaja == 'Morelia' and tarima == 'Metal':
            kgNetos = float(kg) - (float(cajas) * 0.217 + 40)

        elif tipoCaja == 'Saltillo' and tarima == 'Madera':
            kgNetos = float(kg) - (float(cajas) * 1.515 + 22)
        elif tipoCaja == 'Saltillo' and tarima == 'Plastico':
            kgNetos = float(kg) - (float(cajas) * 1.515 + 26.5)
        elif tipoCaja == 'Saltillo' and tarima == 'Metal':
            kgNetos = float(kg) - (float(cajas) * 1.515 + 40)

        elif tipoCaja == 'Bote' and tarima == 'Madera':
            kgNetos = float(kg) - (float(cajas) * 0.77 + 22)
        elif tipoCaja == 'Bote' and tarima == 'Plastico':
            kgNetos = float(kg) - (float(cajas) * 0.77 + 26.5)
        elif tipoCaja == 'Bote' and tarima == 'Metal':
            kgNetos = float(kg) - (float(cajas) * 0.77 + 40)
        elif tipoCaja == 'caja30lb':
            kgNetos = float(cajas) * 13.608 
        elif tipoCaja == 'caja20lb':
            kgNetos = float(cajas) * 9.07
        elif tipoCaja == 'caja10lb':
            kgNetos = float(cajas) * 4.54
        elif tipoCaja == 'caja20kg':
            kgNetos = float(cajas) * 20 
        elif tipoCaja == 'caja12kg':
            kgNetos = float(cajas) * 12 
        elif tipoCaja == 'caja10kg':
            kgNetos = float(cajas) * 10
        elif tipoCaja == 'cubeta12kg':
            kgNetos = float(cajas) * 12 
        elif tipoCaja == '6.5lb':
            kgNetos = float(cajas) * 2.95 
        elif tipoCaja == 'cajaP24':
            kgNetos = float(cajas) * 24
        
        maquila_registro = congeladoP.objects.create(
            pro =pro,
            nombre=nombre,
            kg=kg,
            cajas=cajas,
            tipoCaja=tipoCaja,
            tarima=tarima,
            kgNetos=kgNetos,
            horaEnvio=horaEnvio,
            kgDesp=kgDesp,
            status=status,
            cliente=cliente,
            usuario=request.user,
        )
        # establecer fechaOriginal solo al crear un nuevo registro
        if status == 'entrada':
            maquila_registro.fechaOriginal = horaEnvio
            maquila_registro.save()
            messages.error(request, f'Movimiento {Con} registrado')


    else:
        messages.error(request, 'Movimiento invalido')
    return redirect('/coPro')
@user_passes_test(is_admin, login_url='/error/' )
def eliminarCongeladoP(request, fruta_id):
    Fruta = congeladoP.objects.get(Con=fruta_id)
    Fruta.delete()
    messages.success(request, '¡Movimiento Eliminado!')

    return redirect('/coPro')
@user_passes_test(is_admin, login_url='/error/' )
def edicionCongeladoP(request, fruta_id):
    fruta_obj = congeladoP.objects.get(Con=fruta_id)

    if request.method == 'POST':
        Con = request.POST.get('Con')
        nombre = request.POST.get('nombre')
        unidades = request.POST.get('unidades')
        tipoCaja = request.POST.get('tipoCaja')
        kg = request.POST.get('kg')
        status = request.POST.get('status')
        hora_llegada = request.POST.get('hora_llegada')
        procesos = request.POST.get('procesos')
        horas_envio = request.POST.get('horas_envio')
        fecha = request.POST.get('fecha')
        cliente= request.POST.get('cliente')

        fruta_obj.Con = Con
        fruta_obj.nombre = nombre
        fruta_obj.unidades = unidades
        fruta_obj.tipoCaja = tipoCaja
        fruta_obj.kg = kg
        fruta_obj.status = status
        fruta_obj.hora_llegada = hora_llegada
        fruta_obj.procesos = procesos
        fruta_obj.horas_envio = horas_envio
        fruta_obj.fecha = fecha
        fruta_obj.cliente = cliente
        fruta_obj.save()

        messages.success(request, '¡Movimiento Actualizado!')
        return redirect('/coPro')

    return render(request, 'edicionCongeladoP.html', {'fruta': fruta_obj})
@user_passes_test(is_admin, login_url='/error/' )
#--------------------------------------reporte en excel de produccion-------------------------------
def reporte(request):
    # Ruta del archivo
    archivo_existente = "Aplicaciones/Trabajo/FORMATO DE SEGUIMIENTO.xlsx"
    # Cargar el libro de trabajo existente
    workbook = load_workbook(archivo_existente)

    # Obtener o crear las hojas necesarias
    hojas_necesarias = {
        'Almacen': ['maquila_origen','movimiento', 'proveedor__nombre','nombre','unidades','tipoCaja','tarima','kg','kgNetos','fecha','hora','maq','usuario__username','maquila'],
        'Maquila': ['id', 'lote_id', 'nombre', 'maquila', 'cantidad', 'cajas', 'tipoCaja', 'tarima', 'kgNetos', 'horaEnvio', 'hora', 'usuario__username'],
        'Lavado': ['id','lote_id','nombre','kg','cajas','tipoCaja','tarima','kgNetos','kgDesp','horaEnvio','hora','usuario__username'],
        'Congelado': ['id','lote_id','nombre','kg','cajas','tipoCaja','tarima','kgNetos','kgDesp','horaEnvio','hora','usuario__username'],
        'Quebrado': ['id','lote_id','nombre','kg','cajas','tipoCaja','tarima','kgNetos','kgDesp','horaEnvio','hora','usuario__username'],
        'CongeladoP': ['pro','nombre','kg','cajas','tipoCaja','tarima','kgNetos','kgDesp','horaEnvio','status','usuario__username'],
        'Entrada de cajas': ['folio','proveedor','tipos','cantidad','fecha'],
        'Salida de cajas': ['movimiento','folio','proveedor','tipos','cantidad_salida','fecha','usuario','hora'],
        'Concervacion': ['lote','tarima','aceptado','tipoCaja','nombre','cajas','kgT','fecha','hora','cliente']
    }

    for nombre_hoja, columnas in hojas_necesarias.items():
        # Obtener o crear la hoja
        hoja = workbook[nombre_hoja] if nombre_hoja in workbook.sheetnames else workbook.create_sheet(title=nombre_hoja)

        # Si la hoja está recién creada, agregar los encabezados
        if hoja.max_row == 0:
            hoja.append(columnas)

        # Obtener los datos de la tabla correspondiente y agregarlos a la hoja
        if nombre_hoja == 'Almacen':
            fruta_data = fruta.objects.values_list(*columnas).order_by('movimiento')
        elif nombre_hoja == 'Maquila':
            fruta_data = Maquila.objects.values_list(*columnas).order_by('id')
        elif nombre_hoja == 'Lavado':
            fruta_data = lavado.objects.values_list(*columnas).order_by('id')
        elif nombre_hoja == 'Congelado':
            fruta_data = congelado.objects.values_list(*columnas).order_by('id')
        elif nombre_hoja == 'Quebrado':
            fruta_data = quebrado.objects.values_list(*columnas).order_by('id')
        elif nombre_hoja == 'CongeladoP':
            fruta_data = congeladoP.objects.values_list(*columnas).order_by('pro')
        elif nombre_hoja == 'Entrada de cajas':
            fruta_data = cajas.objects.values_list(*columnas).order_by('folio')
        elif nombre_hoja == 'Salida de cajas':
            fruta_data = cajasSalida.objects.values_list(*columnas).order_by('movimiento')
        elif nombre_hoja == 'Concervacion':
            fruta_data = concervacion.objects.values_list(*columnas).order_by('lote')

        for row in fruta_data:
            # Convertir las fechas y horas si es necesario
            fecha_indices = [i for i, col in enumerate(columnas) if col == 'fecha']
            hora_indices = [i for i, col in enumerate(columnas) if col == 'hora']
            for idx in fecha_indices:
                row = list(row)
                fecha = row[idx]
                if isinstance(fecha, date):
                    row[idx] = fecha.strftime('%d/%m/%Y')
            for idx in hora_indices:
                row = list(row)
                hora = row[idx]
                if isinstance(hora, time):
                    row[idx] = hora.strftime('%H:%M:%S')
            hoja.append(row)

    # Configurar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=producción.xlsx'

    # Guardar el libro de trabajo en la respuesta HTTP
    workbook.save(response)

    return response
#-------------------------------------------------notas------------------------------------------
@user_passes_test(is_admin, login_url='/error/' )
def crear_nota(request):
    frutas = fruta.objects.filter(en_nota=False).order_by('-movimiento')
    recepcion = recep.objects.filter(en_nota=False).order_by('-id')
    notas = Nota.objects.filter().order_by('-folio')
    prove = proveedores.objects.all()

    if request.method == 'POST':
        # Obtén el último movimiento
        ultimo_movimiento = Nota.objects.order_by('-folio').first()

        # Calcula el nuevo movimiento
        nuevo_movimiento = 1 if not ultimo_movimiento else ultimo_movimiento.folio + 1
        fecha = request.POST.get('fecha')
        comprobante = request.POST.get('comprobante')
        prov = request.POST.get('prov')
        descuento = Decimal(request.POST.get('descuento', '0'))  # Convierte a Decimal
        precio = Decimal(request.POST.get('precio', '0'))  # Convierte a Decimal
        movimientos_seleccionados = request.POST.getlist('movimientos_seleccionados')
        movimientos_seleccionados2 = request.POST.getlist('movimientos_seleccionados2')
        # Convertir los valores a enteros
        movimientos_str = [str(movimiento) for movimiento in movimientos_seleccionados]
        recep_str = [str(recep_id) for recep_id in movimientos_seleccionados2]

        # Convertir los movimientos a objetos Movimiento
        movimientos_objetos = Nota.objects.filter(movi__in=movimientos_seleccionados)
        # Filtrar solo las frutas seleccionadas
        registros = fruta.objects.filter(movimiento__in=movimientos_seleccionados)
        movi = movimientos_seleccionados

        total_kilos = registros.aggregate(total_kilos=Sum('kgNetos'))['total_kilos'] or 0
        total_cajas = registros.aggregate(total_cajas=Sum('unidades'))['total_cajas'] or 0
        tarima = registros.aggregate(total_cajas=Sum('tarima'))['total_cajas'] or 0
        total_kilosB = registros.aggregate(total_kilosB=Sum('kg'))['total_kilosB'] or 0
        total_kilosN = registros.aggregate(total_kilosN=Sum('kgNetos'))['total_kilosN'] or 0
        total_kilosC = registros.aggregate(total_kilosC=Sum('unidades'))['total_kilosC'] or 0

        primer_movimiento = fruta.objects.filter(movimiento__in=movimientos_seleccionados).first()

        # Obtener el nombre del primer movimiento
        nombre_primer_movimiento = primer_movimiento.nombre if primer_movimiento else None



        totalC = 0
        totalT = 0
        kilosB = total_kilosB
        kilosN = total_kilosN
        unidades = total_cajas
        proveedores_seleccionados = set()
        frutas_seleccionados = set()

        for movimiento in movimientos_seleccionados:
            fruta_obj = frutas.filter(movimiento=movimiento).first()
            if fruta_obj:
                proveedor_nombre = fruta_obj.proveedor.nombre
                proveedores_seleccionados.add(proveedor_nombre)
        primer_proveedor = proveedores_seleccionados.pop()
        proveedor=primer_proveedor

        for movimiento in movimientos_seleccionados:
            fruta_obj = frutas.filter(movimiento=movimiento).first()
            if fruta_obj:
                fruta_nombre = fruta_obj.nombre.nombre
                frutas_seleccionados.add(fruta_nombre)
        primer_fruta = frutas_seleccionados.pop()
        nombre=primer_fruta

        for registro in registros:
            tipoCaja = registro.tipoCaja
            tarima = registro.tarima
            unidades_fruta = registro.unidades 
            
            taraC = 0  # Inicializa tara_cajas aquí
            taraT = 0  # Inicializa tara_tarima aquí
            

            if tipoCaja == 'Proveedora':
                taraC = Decimal(unidades) * Decimal("1.68")
            elif tipoCaja == 'Tijuana':
                taraC = Decimal(unidades) * Decimal ("1.3 ")
            elif tipoCaja == 'Colima':
                taraC = Decimal(unidades) * Decimal ("2.18")
            elif tipoCaja == 'Lima':
                taraC = Decimal(unidades) * Decimal ("0.7")
            elif tipoCaja == 'Jacona':
                taraC = Decimal(unidades) * Decimal ("1")
            elif tipoCaja == 'Regilla':
                taraC = Decimal(unidades) * Decimal ("1.68")
            elif tipoCaja == 'Morelia':
                taraC = Decimal(unidades) * Decimal ("0.217")
            elif tipoCaja == 'Saltillo':
                taraC = Decimal(unidades) * Decimal ("1.515")
            elif tipoCaja == 'Bote':
                taraC = Decimal(unidades) * Decimal ("0.77")

            if tarima == 'Madera':
                taraT =  22
            elif tarima == 'Plastico':
                taraT = 26.5
            elif tarima == 'Metal':
                taraT =  40 
                 

            totalC += taraC
            totalT += taraT
            registro.tara_cajas_calculada = taraC
            registro.tara_tarima_calculada = taraT

        # Calcular el descuento sobre el total de kilos
        descuento_total = ((total_kilos * descuento) / 100)

        # Calcular el total a pagar restando el descuento
        total_pagar = total_kilos - descuento_total

        totalP = precio * total_pagar
        
        # Supongamos que estás obteniendo un conjunto de consultas como esto
        usuarios = fruta.objects.filter(movimiento__in=movimientos_seleccionados)

        # Asegúrate de que haya al menos un objeto en el conjunto de consultas
        if usuarios.exists():
            # Obtén el primer objeto del conjunto
            primer_usuario = usuarios.first()
            
            # Accede al atributo 'usuario' del objeto
            usuario = primer_usuario.usuario
            # Hacer algo con 'usuario' aquí
        else:
            # Manejar el caso en que no hay usuarios encontrados
            usuario = None  # o manejar de alguna otra manera

        prov = proveedores.objects.get(id = prov)
        # Crear la nota
        nueva_nota = Nota.objects.create(
            folio=nuevo_movimiento,  # Ajusta el campo adecuadamente
            descuento=descuento,
            precio=precio,
            totalK=total_pagar,  # Usar el total calculado después del descuento
            totalP=totalP,
            totalC=totalC,
            totalT=totalT,
            taraC=taraC,
            taraT=taraT,
            fecha = fecha,
            kilosB=kilosB,
            kilosN = kilosN,
            unidades=unidades,
            proveedor=proveedor,
            movi=','.join(movimientos_str) ,
            movi2=','.join(recep_str) ,
            nombre = nombre,
            usuario = usuario,
            prov=prov
        )
        # Cuando necesitas trabajar con los números de movimientos
        numeros_de_movimientos = [int(num) for num in nueva_nota.movi.split(',') if num]
        numeros_de_movimientos2 = [int(num) for num in nueva_nota.movi2.split(',') if num]

        # Obtener los movimientos seleccionados
        movimientos_seleccionados = request.POST.getlist('movimientos_seleccionados')
        movimientos_seleccionados2 = request.POST.getlist('movimientos_seleccionados2')

        # Actualizar las frutas seleccionadas con la referencia a la nota
        frutas_seleccionadas = fruta.objects.filter(movimiento__in=movimientos_seleccionados)
        frutas_seleccionadas2 = recep.objects.filter(id__in=movimientos_seleccionados2)
        frutas_seleccionadas.update(en_nota=True, nota=nueva_nota)
        frutas_seleccionadas2.update(en_nota=True, nota=nueva_nota)

        return redirect('crear_nota')  # Ajusta 'ruta_a_redirigir' según tu configuración
    
    return render(request, 'crear_nota.html', {'Frutas': frutas, 'proveedores2': prove,'notas':notas, 'recepcion':recepcion})

@user_passes_test(is_admin, login_url='/error/' )
def edicionNota(request, nota_id):
    frutas = fruta.objects.filter(en_nota=False).order_by('-movimiento')
    notas = get_object_or_404(Nota, folio=nota_id)
    prov = proveedores.objects.all()

    if request.method == 'POST':
        fecha = request.POST.get('fecha')
        comprobante = request.POST.get('comprobante')
        descuento = Decimal(request.POST.get('descuento', '0'))
        precio = Decimal(request.POST.get('precio', '0'))

        # Restablecer el valor inicial antes de aplicar el descuento
        total_inicial = notas.totalK

        # Calcular el descuento sobre kilosN
        descuento_total = ((notas.kilosN * descuento) / 100)

        # Calcular el total a pagar restando el descuento
        total_pagar = notas.kilosN - descuento_total

        totalP = precio * total_pagar

        # Agregar la lógica para manejar la imagen del comprobante
        try:
            comprobante_file = request.FILES['comprobante']
        except MultiValueDictKeyError:
            comprobante_file = None

        print("Descuento:", descuento)
        print("Precio:", precio)
        print("total_pagar:", total_pagar)
        print("descuento_total:", descuento_total)
        print("totalP:", totalP)
     

        
        # Actualizar la nota existente con los nuevos valores
        notas.descuento = descuento
        notas.precio = precio
        notas.totalK = total_pagar
        notas.totalP = totalP

        if comprobante_file:
            notas.comprobante = comprobante_file

        # Guardar los cambios en la nota
        notas.save()

        # Redirige a la página de detalle de la nota o a donde sea necesario
        return redirect('crear_nota')  # Ajusta según tu configuración

    return render(request, 'edicionNota.html', {'nota': notas, 'Frutas': frutas, 'proveedores': prov})

@user_passes_test(is_admin, login_url='/error/' )
def excelNota(request,folio_id):
    
    # Filtrar los registros de fruta para el movimiento específico
    frutas = fruta.objects.filter()
    notas = Nota.objects.filter(folio=folio_id)
    # Obtener el objeto Nota específico para el folio proporcionado
    try:
        nota = Nota.objects.get(folio=folio_id)
    except Nota.DoesNotExist:
        return HttpResponse('No se encontró la nota con folio {}'.format(folio_id))

     # Obtener el objeto Nota específico para el folio proporcionado
    if not notas.exists():
        # Manejar el caso en el que no hay registros para el movimiento dado
        # Puedes redirigir a una página de error o mostrar un mensaje apropiado
        return HttpResponse('No hay registros para el movimiento {}'.format(folio_id))

    # Cargar el archivo Excel preexistente
    workbook = openpyxl.load_workbook('Aplicaciones/Trabajo/VALES.xlsx')
    sheet = workbook.active

    # Obtener la fecha actual
    fecha_actual = datetime.now()
    # Obtener el nombre del usuario desde el objeto request.user
    usuario_nombre = request.user.username if request.user.is_authenticated else "Usuario no autenticado"

    # Formatear la fecha en el formato deseado
    fecha_formateada = fecha_actual.strftime('%Y-%m-%d %H:%M:%S')
    # Asignar los datos a celdas específicas (usando el primer registro)
      # Puedes ajustar esto según tus necesidadesvc
     # Puedes ajustar esto según tus necesidadesvc
    fruta_obj = frutas.first()  # Puedes ajustar esto según tus necesidadesvc
    nota_obj = notas.first()
    datos = f'{fruta_obj.proveedor.nombre},{fruta_obj.maq}'
    

    sheet['B7'] = nota_obj.proveedor
    sheet['B34'] = nota_obj.proveedor
    sheet['B8'] = nota_obj.nombre
    sheet['B35'] = nota_obj.nombre
    sheet['K2'] = nota_obj.folio
    sheet['K29'] = nota_obj.folio
    sheet['K3'] = nota_obj.fecha
    sheet['K30'] = nota_obj.fecha
    sheet['K24'] = nota_obj.usuario
    sheet['K51'] = nota_obj.usuario
    sheet['A11'] = nota_obj.unidades
    sheet['A38'] = nota_obj.unidades
    sheet['C11'] = nota_obj.kilosB
    sheet['C38'] = nota_obj.kilosB
    sheet['E11'] = nota_obj.kilosN
    sheet['E38'] = nota_obj.kilosN
    sheet['H11'] = nota_obj.descuento
    sheet['H38'] = nota_obj.descuento
    sheet['K11'] = nota_obj.totalK
    sheet['K38'] = nota_obj.totalK
    sheet['C22'] = nota_obj.precio
    sheet['C49'] = nota_obj.precio
    sheet['C24'] = nota_obj.totalP
    sheet['C51'] = nota_obj.totalP
    sheet['C19'] = nota_obj.movi
    sheet['C46'] = nota_obj.movi


    # Establecer propiedades de protección para evitar la edición del archivo
    for row in sheet.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    # Proteger la hoja
    sheet.protection.sheet = True

    # Guardar el archivo Excel modificado y enviarlo como respuesta
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Folio{}.xlsx'.format(folio_id)

    return response

#----------------------------------------------CALIDAD-----------------------------------------------
#Recepcion
@user_passes_test(is_admin_cal, login_url='/error/' )
def muestreo(request):
    if request.method == 'POST':
        fruta = request.POST.get('fruta')  # Convertir a entero
        usuario = request.user
        fecha = request.POST.get('fecha')
        hora = request.POST.get('hora')
        productor_id = int(request.POST.get('productor'))  # Convertir a entero
        brix = Decimal(request.POST.get('brix'))
        ph = Decimal(request.POST.get('ph'))
        muestra = Decimal(request.POST.get('muestra'))
        tipoCaja = request.POST.get('tipoCaja')
        mancha = Decimal(request.POST.get('mancha'))
        tamano = Decimal(request.POST.get('tamano'))
        caracter = Decimal(request.POST.get('caracter'))
        sMadura = Decimal(request.POST.get('sMadura'))
        defecto = Decimal(request.POST.get('defecto'))
        verde = Decimal(request.POST.get('verde'))
        podrida = Decimal(request.POST.get('podrida'))
        lodo = Decimal(request.POST.get('lodo'))
        larva = request.POST.get('larva')
        mExtrana = request.POST.get('mExtrana')
        limpieza = request.POST.get('limpieza')
        aQuimicos = request.POST.get('aQuimicos')
        aPlagas = request.POST.get('aPlagas')
        tDefecto = request.POST.get('tDefecto')
        muestreo = request.FILES.get('muestreo')  # Usar get para archivos
        variedad = request.POST.get('variedad')

        productor = proveedores.objects.get(id=productor_id)
        fruta = materia.objects.get(id=fruta)

        tDefecto = (tamano + caracter + sMadura + defecto + verde + podrida + lodo + mancha)
        
        if fruta.id == 1:
            porT = ((tamano * Decimal('100')) / muestra) 
            if porT > 5: 
                porT -= 5  # Restar 5 y asignar el resultado nuevamente a porT
            else:
                porT = 0
            porC = ((caracter * Decimal('100')) / muestra) 
            if porC > 5: 
                porC -= 5
            else:
                porC = 0
            porS = ((sMadura * Decimal('100')) / muestra) 
            if porS > 5: 
                porS -= 5
            else:
                porS = 0
            porD = ((defecto * Decimal('100')) / muestra) 
            if porD > 5: 
                porD -= 5
            else:
                porD = 0
            porV = ((verde * Decimal('100')) / muestra) 
            if porV > 7: 
                porV -= 7
            else:
                porV = 0
            porP = ((podrida * Decimal('100')) / muestra) 
            if porP > 5: 
                porP -= 5
            else:
                porP = 0

            porL = ((lodo * Decimal('100')) / muestra) 
            if porL > 5: 
                porL -= 5
            else:
                porL = 0
            porM = ((mancha * Decimal('100')) / muestra) 
            if porM > 5: 
                porM -= 5
            else:
                porM = 0


            print(porT , porC , porS , porD , porV , porP , porL , porM)
            descuento = ((porT + porC + porS + porD + porV + porP + porL + porM))
            
            print("Valor de fruta:", fruta)
            print("Total de descuento:", descuento)

        elif fruta.id == 2:
            porT = ((tamano * Decimal('100')) / muestra) 
            if porT > 5: 
                porT -= 5  # Restar 5 y asignar el resultado nuevamente a porT
            else:
                porT = 0
            porC = ((caracter * Decimal('100')) / muestra) 
            if porC > 10: 
                porC -= 10
            else:
                porC = 0
            porS = ((sMadura * Decimal('100')) / muestra) 
            if porS > 5: 
                porS -= 5
            else:
                porS = 0
            porD = ((defecto * Decimal('100')) / muestra) 
            if porD > 5: 
                porD -= 5
            else:
                porD = 0
            porV = ((verde * Decimal('100')) / muestra) 
            if porV > 5: 
                porV -= 5
            else:
                porV = 0
            porP = ((podrida * Decimal('100')) / muestra) 
            if porP > 5: 
                porP -= 5
            else:
                porP = 0

            porL = ((lodo * Decimal('100')) / muestra) 
            if porL > 5: 
                porL -= 5
            else:
                porL = 0
            porM = ((mancha * Decimal('100')) / muestra) 
            if porM > 5: 
                porM -= 5
            else:
                porM = 0


            print(porT , porC , porS , porD , porV , porP , porL , porM)
            descuento = ((porT + porC + porS + porD + porV + porP + porL + porM))
            
            print("Valor de fruta:", fruta)
            print("Total de descuento:", descuento)
        elif fruta.id == 3:
            porT = ((tamano * Decimal('100')) / muestra) 
            if porT > 5: 
                porT -= 5  # Restar 5 y asignar el resultado nuevamente a porT
            else:
                porT = 0
            porC = ((caracter * Decimal('100')) / muestra) 
            if porC > 10: 
                porC -= 10
            else:
                porC = 0
            porS = ((sMadura * Decimal('100')) / muestra) 
            if porS > 5: 
                porS -= 5
            else:
                porS = 0
            porD = ((defecto * Decimal('100')) / muestra) 
            if porD > 5: 
                porD -= 5
            else:
                porD = 0
            porV = ((verde * Decimal('100')) / muestra) 
            if porV > 5: 
                porV -= 5
            else:
                porV = 0
            porP = ((podrida * Decimal('100')) / muestra) 
            if porP > 5: 
                porP -= 5
            else:
                porP = 0

            porL = ((lodo * Decimal('100')) / muestra) 
            if porL > 5: 
                porL -= 5
            else:
                porL = 0
            porM = ((mancha * Decimal('100')) / muestra) 
            if porM > 5: 
                porM -= 5
            else:
                porM = 0


            print(porT , porC , porS , porD , porV , porP , porL , porM)
            descuento = ((porT + porC + porS + porD + porV + porP + porL + porM))
            
            print("Valor de fruta:", fruta)
            print("Total de descuento:", descuento)
        elif fruta.id == 4:
            porT = ((tamano * Decimal('100')) / muestra) 
            if porT > 5: 
                porT -= 5  # Restar 5 y asignar el resultado nuevamente a porT
            else:
                porT = 0
            porC = ((caracter * Decimal('100')) / muestra) 
            if porC > 10: 
                porC -= 10
            else:
                porC = 0
            porS = ((sMadura * Decimal('100')) / muestra) 
            if porS > 5: 
                porS -= 5
            else:
                porS = 0
            porD = ((defecto * Decimal('100')) / muestra) 
            if porD > 5: 
                porD -= 5
            else:
                porD = 0
            porV = ((verde * Decimal('100')) / muestra) 
            if porV > 8: 
                porV -= 8
            else:
                porV = 0
            porP = ((podrida * Decimal('100')) / muestra) 
            if porP > 5: 
                porP -= 5
            else:
                porP = 0

            porL = ((lodo * Decimal('100')) / muestra) 
            if porL > 5: 
                porL -= 5
            else:
                porL = 0
            porM = ((mancha * Decimal('100')) / muestra) 
            if porM > 5: 
                porM -= 5
            else:
                porM = 0


            print(porT , porC , porS , porD , porV , porP , porL , porM)
            descuento = ((porT + porC + porS + porD + porV + porP + porL + porM))
            
            print("Valor de fruta:", fruta)
            print("Total de descuento:", descuento)
        muest = recep.objects.create(
            fruta=fruta,
            usuario=usuario,
            fecha=fecha,
            hora=hora,
            productor=productor,
            brix=brix,
            ph=ph,
            muestra=muestra,
            tipoCaja=tipoCaja,
            mancha=mancha,
            tamano=tamano,
            caracter=caracter,
            sMadura=sMadura,
            defecto=defecto,
            verde=verde,
            podrida=podrida,
            lodo=lodo,
            larva=larva,
            mExtrana=mExtrana,
            limpieza=limpieza,
            aQuimicos=aQuimicos,
            aPlagas=aPlagas,
            tDefecto=tDefecto,
            descuento=descuento,
            muestreo=muestreo,
            variedad=variedad
        )
        muest.save()
        messages.success(request, '¡Movimiento Guardado!')
        return redirect('/mues3')
    return redirect('/mues3')

@user_passes_test(is_admin_cal, login_url='/error/' )
def mues(request):
    frutas = recep.objects.all()
    prov = proveedores.objects.all()
    frut = materia.objects.all()
    return render(request, 'recep.html', {'Frutas': frutas,'prov':prov,'frut':frut})
@user_passes_test(is_admin_cal, login_url='/error/' )
def mues2(request):
    frutas = recep.objects.all().order_by('-id')
    return render(request, 'recep2.html', {'Frutas': frutas})
@user_passes_test(is_admin_cal, login_url='/error/' )
def edicionMuestreos(request,mues_id):
    mues =recep.objects.get(id=mues_id)
    prov =proveedores.objects.all()
    if request.method == 'POST':
        fruta = request.POST.get('fruta')
        usuario = request.POST.get('usuario')
        fecha = request.POST.get('fecha')
        hora = request.POST.get('hora')
        productor = request.POST.get('productor')
        brix = request.POST.get('brix')
        ph = request.POST.get('ph')
        muestra = request.POST.get('muestra')
        tipoCaja = request.POST.get('tipoCaja')
        mancha = request.POST.get('mancha')
        tamano = request.POST.get('tamano')
        caracter = request.POST.get('caracter')
        sMadura = request.POST.get('sMadura')
        defecto = request.POST.get('defecto')
        verde = request.POST.get('verde')
        podrida = request.POST.get('podrida')
        lodo = request.POST.get('lodo')
        larva = request.POST.get('larva')
        mExtrana= request.POST.get('mExtrana')
        limpieza = request.POST.get('limpieza')
        aQuimicos = request.POST.get('aQuimicos')
        aPlagas = request.POST.get('aPlagas')
        tDefecto = request.POST.get('tDefecto')
        variedad = request.POST.get('variedad')
        descuento = request.POST.get('descuento')
        muestreo = request.POST.get('muestreo')
        usuario = request.user
        productor = proveedores.objects.get(id=productor)
        # Agregar la lógica para manejar la imagen del comprobante
        try:
            muestreo_file = request.FILES['muestreo']
        except MultiValueDictKeyError:
            muestreo_file = None

        
        # Guardar los cambios en la nota
        mues.fruta = fruta
        mues.usuario = usuario
        mues.fecha = fecha
        mues.hora = hora 
        mues.productor = productor
        mues.brix = brix
        mues.ph = ph
        mues.muestra = muestra
        mues.tipoCaja = tipoCaja
        mues.mancha = mancha
        mues.tamano = tamano
        mues.caracter = caracter
        mues.sMadura = sMadura
        mues.defecto = defecto
        mues.verde = verde
        mues.podrida = podrida
        mues.lodo = lodo
        mues.larva = larva
        mues.mExtrana = mExtrana
        mues.limpieza = limpieza
        mues.aQuimicos = aQuimicos
        mues.aPlagas = aPlagas
        mues.tDefecto = tDefecto
        mues.variedad = variedad
        mues.descuento = descuento
        mues.muestreo = muestreo
        if muestreo_file:
            mues.muestreo = muestreo_file
        mues.save()
 
        messages.success(request, '!Actualizado!')
        return redirect('/mues2')
    return render(request, 'edicionMuestreo.html', {'mues':mues,'prov':prov})
@user_passes_test(is_admin_cal, login_url='/error/' )
def eliminarMuestreos(request,mues_id):
    Fruta = recep.objects.get(id=mues_id)
    Fruta.delete()
    messages.success(request, '¡Movimiento Eliminado!')

    return redirect('/mues2')
@user_passes_test(is_admin_cal, login_url='/error/' )
#Kilos Liberados
def kilosLiberados(request):
    if request.method == 'POST':
        fecha = request.POST.get('fecha')  # Convertir a entero
        usuario = request.user
        hora = request.POST.get('hora')
        cliente = request.POST.get('cliente')
        productos = request.POST.get('productos')
        lote = request.POST.get('lote')
        nTarima = request.POST.get('nTarima')
        loTarima =  request.POST.get('loTarima')
        cantidad = request.POST.get('cantidad')
        kg = request.POST.get('kg')
        kgNetos = request.POST.get('kgNetos')

        productos = materia.objects.get(id=productos)

        fis = fisicoQ.objects.create(
            fecha=fecha,
            usuario=usuario,
            hora=hora,
            cliente=cliente,
            productos=productos,
            lote=lote,
            nTarima=nTarima,
            loTarima=loTarima,
            cantidad=cantidad,
            kg=kg,
            kgNetos=kgNetos,
            )
        fis.save()
        messages.success(request, '¡Movimiento Guardado!')
        return redirect('/kLib2')
    return redirect('/kLib2')
@user_passes_test(is_admin_cal, login_url='/error/' )
def kLib(request):
    fis = fisicoQ.objects.all().order_by('-id')
    producs = materia.objects.all()
    return render(request, 'fQuimicos.html', {'fis':fis,'producs':producs})
def imprimir(request):
    fis = fisicoQ.objects.all()
    if request.method == 'POST':
        # Obtener los movimientos seleccionados
        movimientos_seleccionados = request.POST.getlist('movimientos_seleccionados')

        # Obtener los datos de los movimientos seleccionados
        movimientos = []
        total_cantidad = 0
        total_kg = 0
        total_kg_netos = 0

        for movimiento_id in movimientos_seleccionados:
            movimiento = fisicoQ.objects.get(id=movimiento_id)
            movimientos.append(movimiento)
            total_cantidad += movimiento.cantidad
            total_kg += movimiento.kg
            total_kg_netos += movimiento.kgNetos

        # Datos adicionales para el encabezado
        fecha = movimientos[0].fecha
        cliente = movimientos[0].cliente
        productos = movimientos[0].productos.nombre
        lote = movimientos[0].lote
        usuario = movimientos[0].usuario

        # Renderizar la plantilla de impresión con los datos
        return render(request, 'imprimir.html', {
            'fecha': fecha,
            'cliente': cliente,
            'productos': productos,
            'lote': lote,
            'fis_list': movimientos,
            'total_cantidad': total_cantidad,
            'total_kg': total_kg,
            'total_kg_netos': total_kg_netos,
            'usuario':usuario
        })
    return render(request, 'imprimir.html', {'fis':fis})
#---------------------CATALOGOS---------------------------------------------
#embalaje
@user_passes_test(is_admin, login_url='/error/' )
def agregarEmbalaje(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        peso = request.POST.get('peso')

        emb = embalaje.objects.create(
            nombre = nombre,
            peso = peso
        )
        emb.save()
        messages.success(request, '¡Movimiento Guardado!')
        return redirect('/embase')
@user_passes_test(is_admin, login_url='/error/' )
def embase(request):
    emba = embalaje.objects.all()
    return render(request,'cEmbalaje.html',{'emba':emba})
@user_passes_test(is_admin, login_url='/error/' )
def edicionEmbalaje(request, embalaje_id):
    mues =embalaje.objects.get(id=embalaje_id)
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        peso = request.POST.get('peso')

        mues.nombre = nombre
        mues.peso = peso
        mues.save()
        return redirect('embase')  # Ajusta según tu configuración

    return render(request, 'edicionEmbalaje.html',{'mues':mues})
@user_passes_test(is_admin, login_url='/error/' )
def eliminarEmbalaje(request,embalaje_id):
    emba = embalaje.objects.get(id=embalaje_id)
    emba.delete()
    messages.success(request, '¡Movimiento Eliminado!')

    return redirect('/embase')
#paleta
@user_passes_test(is_admin, login_url='/error/' )
def agregarPaleta(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        peso = request.POST.get('peso')

        emb = paleta.objects.create(
            nombre = nombre,
            peso = peso
        )
        emb.save()
        messages.success(request, '¡Movimiento Guardado!')
        return redirect('/palet')
@user_passes_test(is_admin, login_url='/error/' )
def palet(request):
    emba = paleta.objects.all()
    return render(request,'cPaleta.html',{'emba':emba})
@user_passes_test(is_admin, login_url='/error/' )
def edicionPaleta(request, paleta_id):
    mues = paleta.objects.get(id=paleta_id)
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        peso = request.POST.get('peso')

        mues.nombre = nombre
        mues.peso = peso
        mues.save()
        return redirect('palet')  # Ajusta según tu configuración

    return render(request, 'edicionPaleta.html',{'mues':mues})
@user_passes_test(is_admin, login_url='/error/' )
def eliminarPaleta(request,paleta_id):
    emba = paleta.objects.get(id=paleta_id)
    emba.delete()
    messages.success(request, '¡Movimiento Eliminado!')

    return redirect('/palet')
#catalogos de frutas
@user_passes_test(is_admin, login_url='/error/' )
def agregarBerries(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        codigo = request.POST.get('codigo')
        catalogo = request.POST.get('catalogo')

        emb = materia.objects.create(
            nombre = nombre,
            codigo = codigo,
            catalogo=catalogo
        )
        emb.save()
        messages.success(request, '¡Movimiento Guardado!')
        return redirect('/nombres')
@user_passes_test(is_admin, login_url='/error/' )
def nombres(request):
    nom = materia.objects.all()
    return render(request,'catalogo.html',{'nom':nom})
@user_passes_test(is_admin, login_url='/error/' )
def edicionBerries(request,berries_id):
    mues = materia.objects.get(id=berries_id)
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        codigo = request.POST.get('codigo')
        catalogo = request.POST.get('catalogo')

        mues.nombre = nombre
        mues.codigo = codigo
        mues.catalogo = catalogo
        mues.save()
        return redirect('nombres')  # Ajusta según tu configuración

    return render(request, 'edicionCatalogo.html',{'mues':mues})
@user_passes_test(is_admin, login_url='/error/' )
def eliminarBerries(request,berries_id):
    emba = materia.objects.get(id=berries_id)
    emba.delete()
    messages.success(request, '¡Movimiento Eliminado!')

    return redirect('/nombres')








def registrarDevolucion(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        lote = request.POST.get('lote')
        fruta_instance = fruta.objects.get(movimiento=lote)
        cajas = int(request.POST.get('cajas'))
        tipoCaja = request.POST.get('tipoCaja')
        tarima = request.POST.get('tarima')
        kg = request.POST.get('kg')
        horaEnvio = request.POST.get('horaEnvio')
        kgNetos = request.POST.get('kgNetos')
        kgDesp= request.POST.get('kgDesp')
        hora= request.POST.get('hora')
        unidades_devueltas = int(request.POST.get('unidades_devueltas'))
        kg_devueltos = Decimal(request.POST.get('kg_devueltos'))
        
        nombre = materia.objects.get(id=nombre)
        tipoCaja = embalaje.objects.get(id=tipoCaja)
        tarima = paleta.objects.get(id=tarima)

        kgNetos = Decimal(kg) - ((tipoCaja.peso * cajas) + tarima.peso)

        # Actualizar los campos de fruta
        fruta_instance.kg_procesados += Decimal(kg)
        fruta_instance.unidades_procesadas += cajas
        fruta_instance.kg_devueltos += kg_devueltos
        fruta_instance.unidades_devueltas += unidades_devueltas
        fruta_instance.save()

        # Crear registro de lavado
        devolucion_registro = lavado.objects.create(
            nombre=nombre,
            lote=fruta_instance,
            kg=kg,
            cajas=cajas,
            tipoCaja=tipoCaja,
            tarima=tarima,
            kgNetos=kgNetos,
            horaEnvio=horaEnvio,
            kgDesp=kgDesp,
            usuario=request.user,
            hora=hora,
            unidades_devueltas=unidades_devueltas,
            kg_devueltos=kg_devueltos
        )
        devolucion_registro.save()
        messages.success(request, 'Devolución registrada exitosamente')
        return redirect('/op32')
    
    else:
        messages.error(request, 'Movimiento inválido')
    
    return redirect('op3')

def registrar_lavado(request):
    return render(request, 'registrar_lavado.html')